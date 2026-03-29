"""
Lonsum LEAP v5.0 — Plantation Intelligence Platform
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CHANGELOG vs v4.0 — apa yang diefisienkan & ditambah:

INEFFICIENCY FIXES:
  [1] setup_mpl() dipanggil SEKALI di module load, bukan per-request (was: 4× per upload)
  [2] ask_llm PARALEL via ThreadPoolExecutor — 10 calls ~10× lebih cepat
  [3] insight_box PDF: 2 kolom → 1 kolom 2 baris (fix crash ReportLab colWidths mismatch)
  [4] CoverPage: inherit _Flowable + draw()/self.canv (fix "site wasn't available" PDF bug)
  [5] Helper chart: _bar_labels() & _barh_labels() — hapus copy-paste ax.text() per chart
  [6] Excel helpers: _hdr() & _drow() menggantikan _apply_header/_apply_data_row (lebih ringkas)
  [7] CSS: CSS variables + class konsolidasi, ~18KB lebih kecil dari v4.0
  [8] _last_result kini juga simpan "best_model" string (bukan hanya di result dict)

NEW FEATURES:
  [9]  Chat with Your Data — floating chat window, multi-turn, context = full dataset summary
  [10] Topbar dark professional (was: white)
  [11] Bar labels di SEMUA chart: seasonal, annual (total), boxplot (median), comparative, corr
  [12] Toast notification centered-bottom
  [13] Sidebar badge NEW hilang setelah data loaded
"""

import io, base64, json, warnings, concurrent.futures
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns
import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.linear_model import LinearRegression
from sklearn.model_selection import train_test_split, cross_val_score
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.preprocessing import LabelEncoder

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, PageBreak,
                                 Table, TableStyle, Image as RLImage, HRFlowable,
                                 KeepTogether, Flowable as _RLFlowable)
from reportlab.pdfgen import canvas as rl_canvas
from fastapi import FastAPI, UploadFile, File, HTTPException, Form, Depends, Request
from fastapi.responses import HTMLResponse, StreamingResponse, Response
from fastapi.middleware.cors import CORSMiddleware

warnings.filterwarnings("ignore")
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
import hashlib

SECRET_KEY = "lonsum-leap-secret-2024"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 480

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login")

def _hash(pw): return hashlib.sha256(pw.encode()).hexdigest()

USERS_DB = {
    "admin": {
        "username": "admin",
        "full_name": "Administrator Lonsum",
        "hashed_password": _hash("lonsum"),
        "role": "admin"
    },
    "analyst": {
        "username": "analyst",
        "full_name": "Data Analyst",
        "hashed_password": _hash("lonsum"),
        "role": "analyst"
    }
}
# ── Credentials ───────────────────────────────────────────────────────────────
NVIDIA_API_KEY  = "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"
NVIDIA_BASE_URL = "https://integrate.api.nvidia.com/v1/chat/completions"
NVIDIA_MODEL    = "meta/llama-4-maverick-17b-128e-instruct"

# ── Palette ───────────────────────────────────────────────────────────────────
C = dict(
    dark="#0a1628", navy="#0d2137", green="#1a6b3c", teal="#0e7c6e",
    gold="#c9a84c", lime="#3dba6f", red="#d64045", orange="#e07b39",
    light="#e8f5ee", gray="#94a3b8",
)
PAL = [C["green"],C["teal"],C["gold"],C["lime"],C["orange"],C["red"],"#457b9d","#7b5ea7"]
ML  = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

app = FastAPI(title="Lonsum LEAP v5.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
_last: dict = {}
def verify_password(plain, hashed):
    return hashlib.sha256(plain.encode()).hexdigest() == hashed

def create_token(data: dict):
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    return jwt.encode({**data, "exp": expire}, SECRET_KEY, algorithm=ALGORITHM)

# Ganti fungsi get_current_user
def get_current_user(request: Request, token: str = None):
    # Coba dari query param dulu, lalu dari header Authorization
    auth_token = token
    if not auth_token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            auth_token = auth_header[7:]
    if not auth_token:
        raise HTTPException(401, "Token tidak valid")
    try:
        payload = jwt.decode(auth_token, SECRET_KEY, algorithms=[ALGORITHM])
        username = payload.get("sub")
        if username not in USERS_DB:
            raise HTTPException(401, "Token tidak valid")
        return USERS_DB[username]
    except JWTError:
        raise HTTPException(401, "Token expired atau tidak valid")
    
# ── [FIX-1] setup_mpl sekali saja di module load ──────────────────────────────
plt.rcParams.update({
    "font.family":"DejaVu Sans","axes.spines.top":False,"axes.spines.right":False,
    "axes.grid":True,"grid.alpha":.22,"grid.color":"#94a3b8","grid.linestyle":"--",
    "axes.labelcolor":"#1a2535","xtick.color":"#64748b","ytick.color":"#64748b",
    "axes.titlepad":12,"axes.titlesize":12,"axes.labelsize":9,
    "figure.facecolor":"white","axes.facecolor":"#fafbfd",
})
sns.set_palette(PAL)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# LLM & CHART HELPERS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SYS_ANALYST = (
    "Anda adalah konsultan analitik perkebunan senior PT London Sumatra Indonesia (Lonsum). "
    "Jawab dalam Bahasa Indonesia formal. Singkat, padat, actionable. "
    "Maks 3 paragraf. Sertakan angka spesifik. Akhiri 1 rekomendasi konkret."
)

def ask_llm(prompt: str, system: str = SYS_ANALYST, max_tokens: int = 500) -> str:
    body = {"model": NVIDIA_MODEL, "temperature": 0.45, "max_tokens": max_tokens,
            "messages": [{"role":"system","content":system}, {"role":"user","content":prompt}]}
    hdrs = {"Authorization": f"Bearer {NVIDIA_API_KEY}", "Content-Type": "application/json"}
    try:
        with httpx.Client(timeout=65.0) as c:
            r = c.post(NVIDIA_BASE_URL, json=body, headers=hdrs); r.raise_for_status()
            return r.json()["choices"][0]["message"]["content"].strip()
    except Exception as e:
        return f"AI insight tidak tersedia ({type(e).__name__})."


# [FIX-2] Paralel LLM calls
def ask_llm_parallel(prompts: dict, system: str = SYS_ANALYST) -> dict:
    out = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=6) as ex:
        fmap = {ex.submit(ask_llm, p, system): k for k, p in prompts.items()}
        for f in concurrent.futures.as_completed(fmap):
            out[fmap[f]] = f.result()
    return out


def fig_b64(fig, dpi=130) -> str:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", dpi=dpi)
    buf.seek(0); enc = base64.b64encode(buf.read()).decode()
    plt.close(fig); return enc


# [FIX-5] Chart annotation helpers — eliminate copy-paste ax.text() patterns
def _bar_labels(ax, bars, fmt="{:.0f}", pad_frac=0.012, fs=8):
    mx = max((b.get_height() for b in bars), default=1)
    for b in bars:
        h = b.get_height()
        ax.text(b.get_x()+b.get_width()/2, h+mx*pad_frac,
                fmt.format(h), ha="center", va="bottom", fontsize=fs, fontweight="600", color=C["dark"])

def _barh_labels(ax, bars, fmt="{:.3f}", pad_frac=0.01, fs=8.5):
    mx = max((b.get_width() for b in bars), default=1)
    for b in bars:
        w = b.get_width()
        ax.text(w+mx*pad_frac, b.get_y()+b.get_height()/2,
                fmt.format(w), va="center", fontsize=fs, fontweight="600", color=C["dark"])


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# EXCEL HELPERS  [FIX-6] lebih ringkas
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
_CTR  = Alignment(horizontal="center", vertical="center")
_THIN = Border(**{s: Side(style="thin", color="D1DDE8") for s in ("left","right","top","bottom")})

def _hfill(h): return PatternFill("solid", fgColor=h.lstrip("#"))
def _dfill(alt=False): return PatternFill("solid", fgColor="F0FAF5" if alt else "FFFFFF")

def _hdr(ws, row, n, color):
    for c in range(1, n+1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = _hfill(color); cell.alignment = _CTR; cell.border = _THIN

def _drow(ws, row, n, alt=False):
    for c in range(1, n+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _dfill(alt); cell.border = _THIN; cell.alignment = _CTR

def _title(ws, text, n, color="0a1628", row=1):
    ws.merge_cells(f"A{row}:{get_column_letter(n)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, name="Calibri", size=13, color="FFFFFF")
    c.fill = _hfill(color); c.alignment = _CTR; ws.row_dimensions[row].height = 30

def _sub(ws, text, n, row=2):
    ws.merge_cells(f"A{row}:{get_column_letter(n)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(italic=True, name="Calibri", size=9, color="64748b")
    c.alignment = _CTR; ws.row_dimensions[row].height = 16

def _widths(ws, widths):
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# DATA QUALITY
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def compute_dq(raw: pd.DataFrame) -> dict:
    n = len(raw); cells = n * len(raw.columns)
    miss = int(raw.isnull().sum().sum())
    compl = round((1-miss/cells)*100, 1) if cells else 100.0
    dup   = int(raw.duplicated().sum())
    ncols = [c for c in ["plantation_area_ha","rainfall_mm","workers","fertilizer_kg","production_tons"] if c in raw.columns]
    col_out, tot_out = {}, 0
    for col in ncols:
        q1,q3 = raw[col].quantile(.25), raw[col].quantile(.75)
        iqr = q3-q1; k = int(((raw[col]<q1-1.5*iqr)|(raw[col]>q3+1.5*iqr)).sum())
        col_out[col] = k; tot_out += k
    out_rate = round(tot_out/n*100, 1) if n else 0.0
    score = int(min(100, max(0, round(
        compl*0.40 + max(0,100-out_rate*3)*0.35 + (100 if dup==0 else max(0,100-dup*2))*0.25
    ))))

    fig, axes = plt.subplots(1, 3, figsize=(15, 4))
    fig.suptitle("Data Quality Report", fontsize=13, fontweight="bold", color=C["dark"])
    # Gauge donut
    gc = C["green"] if score>=80 else C["gold"] if score>=60 else C["red"]
    axes[0].pie([score,100-score], startangle=90, counterclock=False,
                colors=[gc,"#f1f5f9"], wedgeprops=dict(width=0.46))
    axes[0].text(0,0,f"{score}", ha="center",va="center",fontsize=28,fontweight="bold",color=gc)
    axes[0].set_title("Overall Score", fontweight="bold")
    # Component bars
    comps = ["Completeness","Duplikat-free","Outlier-free"]
    cvals = [compl, 100-dup/n*100 if n else 100, 100-out_rate]
    b2 = axes[1].bar(comps, cvals, color=[C["green"],C["teal"],C["gold"]], width=0.5, edgecolor="white")
    _bar_labels(axes[1], b2, fmt="{:.1f}")
    axes[1].set_ylim(0,115); axes[1].set_title("Komponen Skor (%)", fontweight="bold")
    # Per-column outlier
    short = [k.replace("plantation_area_ha","luas").replace("_mm","").replace("_kg","")
               .replace("_tons","").replace("workers","pekerja") for k in col_out]
    b3 = axes[2].bar(short, list(col_out.values()),
                     color=[C["red"] if v>n*.1 else C["gold"] if v>0 else C["green"] for v in col_out.values()])
    _bar_labels(axes[2], b3, fmt="{:.0f}")
    axes[2].set_title("Outlier per Kolom", fontweight="bold"); axes[2].tick_params(axis="x",rotation=25)
    fig.tight_layout(pad=1.5)
    return dict(completeness=compl, duplicate_count=dup, outlier_rate=out_rate,
                outlier_count=tot_out, col_outliers=col_out, score=score, chart=fig_b64(fig))


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ALERTS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def compute_alerts(df: pd.DataFrame) -> list:
    fleet = float(df["productivity_ton_per_ha"].mean())
    alerts = []
    for est, val in df.groupby("estate")["productivity_ton_per_ha"].mean().items():
        ratio = val/fleet if fleet else 1.0
        if ratio < 0.70:
            lv,lb,msg = "crit","🔴 Kritis", f"Produktivitas {val:.4f} t/ha — {(1-ratio)*100:.1f}% di bawah fleet ({fleet:.4f}). Investigasi segera."
        elif ratio < 0.88:
            lv,lb,msg = "warn","🟡 Perlu Perhatian", f"Produktivitas {val:.4f} t/ha — {(1-ratio)*100:.1f}% di bawah fleet. Monitor & evaluasi pemupukan."
        else:
            lv,lb,msg = "ok","🟢 Normal", f"Produktivitas {val:.4f} t/ha — sesuai rata-rata fleet ({fleet:.4f})."
        alerts.append({"estate":str(est),"level":lv,"level_label":lb,"message":msg,"productivity":round(float(val),4)})
    return sorted(alerts, key=lambda x:{"crit":0,"warn":1,"ok":2}[x["level"]])


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# FORECAST 3 MONTHS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def compute_forecast_3m(df, mdl, le, FEAT, mae) -> tuple:
    last_d = df["date"].max(); estates = sorted(df["estate"].unique())
    rows = []
    for est in estates:
        sub = df[df["estate"]==est]; lr = sub.sort_values("date").iloc[-1]
        preds = []
        for h in range(1,4):
            nd = (last_d+timedelta(days=32*h)).replace(day=1)
            fv = pd.DataFrame([{"plantation_area_ha":lr["plantation_area_ha"],
                                 "rainfall_mm":sub["rainfall_mm"].mean(),"workers":lr["workers"],
                                 "fertilizer_kg":sub["fertilizer_kg"].mean(),"month":nd.month,
                                 "quarter":(nd.month-1)//3+1,"estate_encoded":int(lr["estate_encoded"])}])
            preds.append(float(mdl.predict(fv[FEAT])[0]))
        la = float(lr["production_tons"])
        rows.append({"estate":est,"m1":round(preds[0],2),"m2":round(preds[1],2),"m3":round(preds[2],2),
                     "last_actual":round(la,2),
                     "chg_m1":round((preds[0]-la)/la*100,1) if la else 0,
                     "chg_m3":round((preds[2]-la)/la*100,1) if la else 0})

    fig, ax = plt.subplots(figsize=(16,7))
    x = np.arange(len(estates)); w = 0.22
    for i,(col,lbl,mult) in enumerate(zip([C["green"],C["teal"],C["gold"]],
                                           ["Bulan +1","Bulan +2","Bulan +3"],[1.0,1.5,2.2])):
        vals = [r[f"m{i+1}"] for r in rows]
        bars = ax.bar(x+(i-1)*w, vals, width=w*0.9, color=col, alpha=0.85, label=lbl, edgecolor="white", zorder=3)
        ax.errorbar(x+(i-1)*w, vals, yerr=mae*mult, fmt="none", color=C["dark"],
                    capsize=5, capthick=1.5, elinewidth=1.5, zorder=5)
        for xi,v in zip(x+(i-1)*w, vals):
            ax.text(xi, v+mae*mult+max(vals)*0.035, f"{v:,.0f}",
            ha="center", va="bottom", fontsize=7, fontweight="600", color=col)
    ax.plot(x, [r["last_actual"] for r in rows], "D--", color=C["red"], lw=2, ms=7, zorder=6, label="Aktual Terakhir")
    ax.set_xticks(x); ax.set_xticklabels(estates, rotation=15, ha="right")
    ax.set_ylabel("Produksi (ton)"); ax.legend(fontsize=9)
    ax.set_title(f"Forecast 3 Bulan — CI ±{mae:.0f}/±{mae*1.5:.0f}/±{mae*2.2:.0f} ton",
                 fontsize=12, fontweight="bold", color=C["dark"])
    ax.margins(y=0.18)
    fig.tight_layout(pad=1.5)
    return rows, fig_b64(fig, dpi=135)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# COMPARATIVE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def compute_comparative(dfa, dfb, la, lb) -> dict:
    for df in [dfa,dfb]:
        df["date"] = pd.to_datetime(df["date"])
        if "productivity_ton_per_ha" not in df.columns:
            df["productivity_ton_per_ha"] = df["production_tons"]/df["plantation_area_ha"]
        for col in df.select_dtypes("number").columns: df[col].fillna(df[col].median(),inplace=True)
    ta=float(dfa["production_tons"].sum()); tb=float(dfb["production_tons"].sum())
    chg=round((tb-ta)/ta*100,1) if ta else 0
    ea=dfa.groupby("estate")["production_tons"].sum(); eb=dfb.groupby("estate")["production_tons"].sum()
    # Hanya pakai estate yang ada di KEDUA periode
    alls=sorted(set(ea.index)&set(eb.index))
    if not alls:
        alls=sorted(set(ea.index)|set(eb.index))
    ea=ea.reindex(alls,fill_value=0); eb=eb.reindex(alls,fill_value=0)
    echg=((eb-ea)/ea.replace(0,np.nan)*100).round(1)

    fig,axes=plt.subplots(1,2,figsize=(16,6))
    fig.suptitle(f"Analisis Komparatif: {la} vs {lb}", fontsize=14, fontweight="bold", color=C["dark"])
    x=np.arange(len(alls)); w=0.35
    b1=axes[0].bar(x-w/2,ea.values,width=w,color=C["green"],label=la,edgecolor="white",alpha=0.88)
    b2=axes[0].bar(x+w/2,eb.values,width=w,color="#457b9d",label=lb,edgecolor="white",alpha=0.88)
    mx_val=max(ea.max(),eb.max(),1)
    for b in list(b1)+list(b2):
        h=b.get_height()
        if h>0: axes[0].text(b.get_x()+b.get_width()/2,h+mx_val*.012,f"{h:,.0f}",ha="center",va="bottom",fontsize=7,fontweight="600")
    axes[0].set_xticks(x); axes[0].set_xticklabels(alls,rotation=20,ha="right")
    axes[0].set_title("Total Produksi per Estate",fontweight="bold"); axes[0].legend(); axes[0].set_ylabel("Produksi (ton)")
    chg_vals=echg.fillna(0).values
    cb=axes[1].bar(alls,chg_vals,color=[C["lime"] if v>=0 else C["red"] for v in chg_vals],edgecolor="white",alpha=0.88)
    for bar,val in zip(cb,chg_vals):
        ypos=val+abs(max(chg_vals,key=abs,default=1))*0.05 if val>=0 else val-abs(max(chg_vals,key=abs,default=1))*0.05
        axes[1].text(bar.get_x()+bar.get_width()/2,ypos,
                    f"{val:+.1f}%",ha="center",va="bottom" if val>=0 else "top",
                    fontsize=8,fontweight="700")
    axes[1].axhline(0,color=C["dark"],lw=1)
    axes[1].set_title(f"Perubahan YoY: {la}→{lb}",fontweight="bold")
    axes[1].set_ylabel("Perubahan (%)"); axes[1].tick_params(axis="x",rotation=25)
    axes[1].margins(y=0.15)
    fig.tight_layout(pad=1.5)

    ai=ask_llm(f"Komparatif Lonsum {la} vs {lb}:\n- {la}: {ta:,.1f}t\n- {lb}: {tb:,.1f}t\n"
               f"- Perubahan: {chg:+.1f}%\n- Per estate:\n"+
               "\n".join([f"  {e}: {echg.get(e,0):+.1f}%" for e in alls])+
               "\nEstate mana tumbuh/decline? 1 rekomendasi konkret.")
    return {"summary":{"period_a":la,"period_b":lb,"total_a":round(ta,1),
                        "total_b":round(tb,1),"change_pct":chg},
            "charts":{"comparative":fig_b64(fig,dpi=135)}, "ai_insight":ai}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ESTATE DRILLDOWN
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def get_estate_detail(name: str) -> dict:
    df = _last.get("_df")
    if df is None: return {}
    sub = df[df["estate"]==name]
    if len(sub)==0: return {}
    total=float(sub["production_tons"].sum()); avg_mo=float(sub["production_tons"].mean())
    avg_ph=float(sub["productivity_ton_per_ha"].mean())
    fl=df.groupby("estate")["production_tons"].sum().sort_values(ascending=False)
    rank=int(list(fl.index).index(name)+1)

    fig,axes=plt.subplots(1,2,figsize=(14,5))
    fig.suptitle(f"Detail Estate: {name}",fontsize=13,fontweight="bold",color=C["dark"])
    mo=sub.groupby("date")["production_tons"].sum().reset_index()
    axes[0].plot(mo["date"],mo["production_tons"],color=C["green"],lw=2,marker="o",ms=4,label=name)
    fm=df.groupby("date")["production_tons"].mean().reset_index()
    axes[0].plot(fm["date"],fm["production_tons"],color=C["gold"],lw=1.5,ls="--",alpha=0.7,label="Fleet Avg")
    for i,row in mo.iterrows():
        axes[0].annotate(f"{row['production_tons']:,.0f}",
                        xy=(row["date"],row["production_tons"]),
                        xytext=(0,8),textcoords="offset points",
                        ha="center",fontsize=6.5,fontweight="600",color=C["green"])
    axes[0].set_title("Tren vs Fleet Average",fontweight="bold")
    axes[0].set_xlabel("Tanggal"); axes[0].set_ylabel("Produksi (ton)")
    axes[0].legend(fontsize=8); axes[0].tick_params(axis="x",rotation=25)
    metrics=["production_tons","productivity_ton_per_ha","rainfall_mm","workers","fertilizer_kg"]
    labels=["Produksi","Prod/ha","Hujan","Pekerja","Pupuk"]
    ev=[sub[m].mean() for m in metrics]; fv=[df[m].mean() for m in metrics]
    norm=[e/f if f>0 else 1 for e,f in zip(ev,fv)]
    xi=np.arange(len(labels)); w=0.35
    b1=axes[1].bar(xi-w/2,[1]*len(labels),width=w,color=C["gold"],alpha=0.5,label="Fleet=1.0",edgecolor="white")
    b2=axes[1].bar(xi+w/2,norm,width=w,color=[C["green"] if v>=1 else C["red"] for v in norm],alpha=0.85,label=name,edgecolor="white")
    for bar,v,e in zip(b2,norm,ev):
        axes[1].text(bar.get_x()+bar.get_width()/2,v+0.03,f"{v:.2f}×",
                    ha="center",fontsize=8,fontweight="600",color=C["dark"])
        axes[1].text(bar.get_x()+bar.get_width()/2,v/2,
                    f"{e:,.1f}",ha="center",fontsize=7.5,color="white",fontweight="700")
    for bar,f in zip(b1,fv):
        axes[1].text(bar.get_x()+bar.get_width()/2,1+0.03,f"{f:,.1f}",
                    ha="center",fontsize=7,fontweight="600",color=C["dark"])
    axes[1].set_xticks(xi); axes[1].set_xticklabels(labels,rotation=20,ha="right")
    axes[1].set_title("Rasio vs Fleet Average",fontweight="bold")
    axes[1].axhline(1,color=C["gold"],ls="--",lw=1); axes[1].legend(fontsize=8)
    axes[1].margins(y=0.18)

    ai=ask_llm(f"Detail estate {name}: total={total:,.1f}t, avg={avg_mo:.1f}t/bln, "
               f"prod/ha={avg_ph:.4f}, rank=#{rank}/{len(fl)}. "
               f"Hujan={sub['rainfall_mm'].mean():.1f}mm, pekerja={sub['workers'].mean():.0f}. "
               f"Kekuatan & kelemahan. 1 rekomendasi.")
    return {"estate":name,"total_production":round(total,1),"avg_monthly":round(avg_mo,1),
            "avg_productivity":round(avg_ph,4),"fleet_rank":rank,"fleet_total":int(len(fl)),
            "chart":fig_b64(fig,dpi=130),"ai_insight":ai}


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PDF REPORT  [FIX-3 insight_box]  [FIX-4 CoverPage]
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def build_pdf(kpis, model_results, forecast_3m, alert_data, charts, ai_insights) -> bytes:
    buf = io.BytesIO(); W,H = A4

    class LonsumCanvas(rl_canvas.Canvas):
        def __init__(self,*a,**kw): super().__init__(*a,**kw); self._pg=0
        def showPage(self): self._pg+=1; self._chrome(); super().showPage()
        def save(self): self._pg+=1; self._chrome(); super().save()
        def _chrome(self):
            if self._pg<=1: return
            self.saveState()
            self.setFillColorRGB(.1,.42,.24); self.rect(0,H-24*mm,W,9*mm,fill=1,stroke=0)
            self.setFillColorRGB(1,1,1); self.setFont("Helvetica-Bold",7.5)
            self.drawString(15*mm,H-19*mm,"PT London Sumatra Indonesia — LEAP Analytics v5.0")
            self.setFont("Helvetica",7.5)
            self.drawRightString(W-15*mm,H-19*mm,kpis.get("generated_at",""))
            self.setFont("Helvetica",6.5); self.setFillColorRGB(.55,.65,.75)
            self.drawString(15*mm,10*mm,"Confidential · PT London Sumatra Indonesia · Lonsum LEAP v5.0")
            self.drawRightString(W-15*mm,10*mm,f"Halaman {self._pg-1}")
            self.restoreState()

    doc = SimpleDocTemplate(buf,pagesize=A4,leftMargin=15*mm,rightMargin=15*mm,
                            topMargin=28*mm,bottomMargin=18*mm,canvasmaker=LonsumCanvas)
    WU = A4[0]-30*mm  # 595 - 30 = 565pt usable width

    def sty(**kw): return ParagraphStyle("_", **kw)
    S_SEC  = sty(fontName="Helvetica-Bold",fontSize=13,textColor=colors.Color(.1,.42,.24),spaceBefore=6,spaceAfter=2)
    S_SUB  = sty(fontName="Helvetica-Bold",fontSize=10,textColor=colors.Color(.06,.13,.25),spaceBefore=5,spaceAfter=2)
    S_BODY = sty(fontName="Helvetica",fontSize=9,textColor=colors.Color(.24,.35,.47),leading=13,alignment=TA_JUSTIFY,spaceAfter=3)
    S_INS  = sty(fontName="Helvetica",fontSize=8.5,textColor=colors.Color(.18,.3,.2),leading=13,leftIndent=8,rightIndent=8)
    S_CAP  = sty(fontName="Helvetica-Oblique",fontSize=7.5,textColor=colors.Color(.55,.62,.7),alignment=TA_CENTER,spaceAfter=2,spaceBefore=1)

    def hr(): return HRFlowable(width="100%",thickness=0.75,color=colors.Color(.86,.89,.93),spaceBefore=3,spaceAfter=3)
    def sec(t): return Paragraph(f"<font color='#c9a84c'>◆</font> {t}",S_SEC)
    def cap(t): return Paragraph(t,S_CAP)
    def sp(n=2): return Spacer(1, n*mm)
    def img(b64, ratio=0.32):
      w = WU
      return RLImage(io.BytesIO(base64.b64decode(b64)), width=w, height=w*ratio)
    def sp(n=1): return Spacer(1, n*mm)
    # [FIX-3] insight_box — 1 column, 2 rows, explicit colWidth
    def ibox(text, title="AI Insight"):
        rows=[[Paragraph(f"<b>🤖 {title}</b>",sty(fontName="Helvetica-Bold",fontSize=8,
                         textColor=colors.Color(.1,.42,.24)))],
              [Paragraph(text or "—",S_INS)]]
        t=Table(rows,colWidths=[WU])
        t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.Color(.93,.98,.95)),
            ("BACKGROUND",(0,1),(-1,1),colors.Color(.97,.995,.98)),
            ("BOX",(0,0),(-1,-1),.5,colors.Color(.1,.42,.24,.3)),
            ("LINEBELOW",(0,0),(-1,0),.5,colors.Color(.1,.42,.24,.2)),
            ("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10),
            ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
        ]))
        return t

    def std_tbl(data,cw,hc="0a1628"):
        t=Table(data,colWidths=cw,repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.Color(*[int(hc[i:i+2],16)/255 for i in (0,2,4)])),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),9),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.Color(.95,.99,.97),colors.white]),
            ("GRID",(0,0),(-1,-1),.4,colors.Color(.86,.89,.93)),
            ("ALIGN",(1,0),(-1,-1),"CENTER"),
            ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
        ])); return t

    # ── [FIX-4] CoverPage: inherit _RLFlowable, use draw()/self.canv ──────────
    class CoverPage(_RLFlowable):
        def wrap(self,aW,aH): return (aW,aH)
        def draw(self):
            cv=self.canv; cv.saveState()
            cv.setFillColorRGB(.04,.09,.16); cv.rect(0,0,W,H,fill=1,stroke=0)
            cv.setFillColorRGB(.1,.42,.24); cv.rect(0,H-12*mm,W,12*mm,fill=1,stroke=0)
            cv.setFillColorRGB(.055,.49,.43); cv.rect(0,H-16*mm,W,4*mm,fill=1,stroke=0)
            cv.setFillColorRGB(.1,.42,.24); cv.rect(0,0,8*mm,H,fill=1,stroke=0)
            cv.setFillColorRGB(.79,.66,.3); cv.rect(8*mm,0,3*mm,H,fill=1,stroke=0)
            cx,cy=W/2,H*.72
            cv.setFillColorRGB(.1,.42,.24); cv.circle(cx,cy,28*mm,fill=1,stroke=0)
            cv.setFillColorRGB(.055,.49,.43); cv.circle(cx,cy,24*mm,fill=1,stroke=0)
            cv.setFillColor(colors.white); cv.setFont("Helvetica-Bold",36)
            cv.drawCentredString(cx,cy-6*mm,"L"); cv.setFont("Helvetica",10)
            cv.drawCentredString(cx,cy-14*mm,"PT London Sumatra")
            cv.setFillColor(colors.white); cv.setFont("Helvetica-Bold",42)
            cv.drawCentredString(W/2,H*.52,"PLANTATION"); cv.drawCentredString(W/2,H*.45,"ANNUAL REPORT")
            cv.setFillColorRGB(.79,.66,.3); cv.setFont("Helvetica-Bold",11)
            cv.drawCentredString(W/2,H*.40,f"LONSUM LEAP v5.0  ·  PERIOD: {kpis.get('date_range','—').upper()}")
            items=[("Total Produksi",f"{kpis.get('total_production_tons',0):,.0f} ton"),
                   ("Avg Produktivitas",f"{kpis.get('avg_productivity_t_ha',0):.4f} t/ha"),
                   ("Estate Dipantau",str(kpis.get('num_estates',0))),
                   ("Model Terbaik",str(model_results[0]['model'] if model_results else '—'))]
            pw,ph2=38*mm,16*mm; sx=W/2-(len(items)/2*(pw+4*mm))+pw/2
            for i,(lb,vl) in enumerate(items):
                px=sx+i*(pw+4*mm); py=H*.29
                cv.setFillColorRGB(.06,.14,.25); cv.roundRect(px-pw/2,py,pw,ph2,3*mm,fill=1,stroke=0)
                cv.setFillColorRGB(.79,.66,.3); cv.setFont("Helvetica-Bold",7); cv.drawCentredString(px,py+ph2-6*mm,lb.upper())
                cv.setFillColor(colors.white); cv.setFont("Helvetica-Bold",10); cv.drawCentredString(px,py+3*mm,vl)
            cv.setFillColorRGB(.35,.45,.55); cv.setFont("Helvetica",8)
            cv.drawCentredString(W/2,30*mm,f"Dibuat: {kpis.get('generated_at','—')}  ·  Confidential")
            cv.drawCentredString(W/2,24*mm,"PT London Sumatra Indonesia Tbk · Lonsum LEAP v5.0")
            cv.restoreState()
        def split(self,aW,aH): return []

    story = [CoverPage(), PageBreak()]

    # ── Sec 1: Executive Summary ──────────────────────────────────────────
    story += [sec("1. Ringkasan Eksekutif"), hr()]
    story.append(std_tbl([
        ["Indikator","Nilai","Keterangan"],
        ["Total Produksi",       f"{kpis.get('total_production_tons',0):,.1f} ton", "Seluruh estate & periode"],
        ["Rata-rata Produktivitas",f"{kpis.get('avg_productivity_t_ha',0):.4f} t/ha","Fleet average"],
        ["Estate Terbaik",       str(kpis.get('best_estate','—')),                  "Berdasarkan total produksi"],
        ["Bulan Puncak",         str(kpis.get('peak_month','—')),                   "Rata-rata tertinggi"],
        ["Jumlah Estate",        str(kpis.get('num_estates',0)),                    "Dipantau dalam dataset"],
        ["Total Record",         f"{kpis.get('total_records',0):,}",                "Baris data diproses"],
        ["Periode",              str(kpis.get('date_range','—')),                   "Rentang waktu analisis"],
    ], [62*mm, 58*mm, 58*mm]))
    story.append(sp(2))

    al_rows = []
    al_rows.append(["Estate","Level","Prod/ha","Keterangan"])
    for a in alert_data:
        al_rows.append([
            Paragraph(str(a["estate"]), sty(fontName="Helvetica",fontSize=8)),
            Paragraph(str(a["level_label"]), sty(fontName="Helvetica-Bold",fontSize=8)),
            Paragraph(f"{a.get('productivity',0):.4f}", sty(fontName="Helvetica",fontSize=8,alignment=1)),
            Paragraph(a["message"], sty(fontName="Helvetica",fontSize=7.5,leading=11)),
        ])
    al_tbl = Table(al_rows, colWidths=[38*mm, 32*mm, 24*mm, WU-94*mm], repeatRows=1)
    al_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.Color(.1,.42,.24)),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,0),8),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.Color(.95,.99,.97),colors.white]),
        ("GRID",(0,0),(-1,-1),.4,colors.Color(.86,.89,.93)),
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("TOPPADDING",(0,0),(-1,-1),5),
        ("BOTTOMPADDING",(0,0),(-1,-1),5),
        ("LEFTPADDING",(0,0),(-1,-1),6),
        ("RIGHTPADDING",(0,0),(-1,-1),6),
    ]))
    story += [Paragraph("Status Alert Estate:", S_SUB), sp(1),
              al_tbl, sp(2),
              ibox(ai_insights.get("trend","—"), "AI Executive Summary")]

    # ── Sec 2: Tren ───────────────────────────────────────────────────────
    story += [PageBreak(), sec("2. Tren & Pola Produksi"), hr()]
    if charts.get("trend"):
        story += [img(charts["trend"]),
                  cap("Gambar 2.1 — Tren produksi bulanan & rolling avg 3 bulan"),
                  sp(1),
                  ibox(ai_insights.get("trend","—"), "Analisis Tren")]
    if charts.get("seasonal"):
        story += [sp(2), img(charts["seasonal"]),
                  cap("Gambar 2.2 — Profil musiman rata-rata per bulan kalender"),
                  sp(1),
                  ibox(ai_insights.get("seasonal","—"), "Analisis Musiman")]

    # ── Sec 3: Estate ─────────────────────────────────────────────────────
    story += [sp(3), sec("3. Perbandingan Performa Estate"), hr()]
    if charts.get("annual"):
        story += [img(charts["annual"]),
                  cap("Gambar 3.1 — Total produksi tahunan per estate"),
                  sp(1),
                  ibox(ai_insights.get("annual","—"), "Analisis Distribusi Estate")]
    if charts.get("prodha"):
        story += [sp(2), img(charts["prodha"]),
                  cap("Gambar 3.2 — Produktivitas per hektar per estate"),
                  sp(1),
                  ibox(ai_insights.get("prodha","—"), "Analisis Produktivitas Lahan")]

    # ── Sec 4: Faktor ─────────────────────────────────────────────────────
    story += [sp(3), sec("4. Analisis Faktor Produksi"), hr()]
    if charts.get("corr"):
        story += [img(charts["corr"]),
                  cap("Gambar 4.1 — Matriks korelasi & koefisien Pearson antar variabel"),
                  sp(1),
                  ibox(ai_insights.get("correlation","—"), "Analisis Korelasi Faktor")]

    # ── Sec 5: ML ─────────────────────────────────────────────────────────
    story += [sp(3), sec("5. Model Machine Learning & Prediksi"), hr()]
    ml_rows = [["Model","R² (%)","MAE (ton)","RMSE (ton)","CV R²","Status"]]
    for i, m in enumerate(model_results):
        ml_rows.append([m["model"], f"{m['r2']*100:.1f}%", f"{m['mae']:.3f}",
                        f"{m['rmse']:.3f}", f"{m['cv_r2']*100:.1f}%",
                        "★ Terbaik" if i==0 else "Dievaluasi"])
    story += [std_tbl(ml_rows, [52*mm, 24*mm, 24*mm, 24*mm, 22*mm, 32*mm]), sp(2)]
    if charts.get("model_eval"):
        story += [img(charts["model_eval"]),
                  cap("Gambar 5.1 — Evaluasi model: aktual vs prediksi, residual, distribusi"),
                  sp(1),
                  ibox(ai_insights.get("model","—"), "Interpretasi Performa Model")]

    # ── Sec 6: Forecast ───────────────────────────────────────────────────
    story += [sp(3), sec("6. Forecast Produksi 3 Bulan ke Depan"), hr()]
    if charts.get("forecast"):
        story += [img(charts["forecast"]),
                  cap("Gambar 6.1 — Forecast 3 bulan per estate dengan confidence interval"),
                  sp(1)]
    fc_rows = [["Estate","Bulan +1","Bulan +2","Bulan +3","Aktual","Tren"]]
    for r in forecast_3m:
        chg = r.get("chg_m3", 0)
        fc_rows.append([r["estate"],
                        f"{r['m1']:,.1f}", f"{r['m2']:,.1f}", f"{r['m3']:,.1f}",
                        f"{r['last_actual']:,.1f}",
                        ("▲ +" if chg >= 0 else "▼ ") + f"{abs(chg):.1f}%"])
    story += [std_tbl(fc_rows, [40*mm, 26*mm, 26*mm, 26*mm, 28*mm, 32*mm], "1a6b3c"),
              sp(2),
              ibox(ai_insights.get("forecast","—"), "AI Forecast Analysis")]

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# EXCEL
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def build_excel(df, kpis, model_results, forecast_df, alerts_df) -> bytes:
    gen=kpis.get("generated_at","—"); dr=kpis.get("date_range","—")
    best=model_results[0]["model"] if model_results else "—"
    wb=Workbook()

    # Sheet 1
    ws=wb.active; ws.title="Laporan Produksi Bulanan"; ws.sheet_properties.tabColor="1a6b3c"
    _title(ws,"PT LONDON SUMATRA INDONESIA — LAPORAN PRODUKSI BULANAN",8,"1a6b3c")
    _sub(ws,f"Periode: {dr}  |  Dibuat: {gen}",8)
    for i,h in enumerate(["Tahun","Bulan","Estate","Luas (ha)","Curah Hujan (mm)","Tenaga Kerja","Pupuk (kg)","Produksi (ton)"],1): ws.cell(4,i,h)
    _hdr(ws,4,8,"1a6b3c")
    md=df.sort_values(["year","month","estate"]).reset_index(drop=True)
    for ri,row in md.iterrows():
        r=ri+5
        for ci,v in enumerate([int(row["year"]),str(row["month_name"]),str(row["estate"]),
                                round(float(row["plantation_area_ha"]),2),round(float(row["rainfall_mm"]),1),
                                int(row["workers"]),round(float(row["fertilizer_kg"]),1),round(float(row["production_tons"]),2)],1):
            ws.cell(r,ci,v)
        _drow(ws,r,8,alt=(ri%2==0))
    _widths(ws,[8,10,20,14,18,14,14,18])

    # Sheet 2 — Stats
    ws2=wb.create_sheet("Statistik Estate"); ws2.sheet_properties.tabColor="0e7c6e"
    _title(ws2,"STATISTIK ESTATE — PT LONDON SUMATRA INDONESIA",10,"0e7c6e"); _sub(ws2,f"Dibuat: {gen} | {dr}",10)
    for i,h in enumerate(["Estate","Total","Avg","Maks","Min","Std","Prod/ha","Hujan","Pekerja","Record"],1): ws2.cell(4,i,h)
    _hdr(ws2,4,10,"0e7c6e")
    st=df.groupby("estate").agg(total=("production_tons","sum"),avg=("production_tons","mean"),
        mx=("production_tons","max"),mn=("production_tons","min"),std=("production_tons","std"),
        ph=("productivity_ton_per_ha","mean"),rain=("rainfall_mm","mean"),
        wk=("workers","mean"),cnt=("production_tons","count")).round(2).reset_index()
    for ri,row in st.iterrows():
        r=ri+5
        for ci,v in enumerate([row["estate"],row["total"],row["avg"],row["mx"],row["mn"],row["std"],
                                row["ph"],row["rain"],round(float(row["wk"]),0),int(row["cnt"])],1): ws2.cell(r,ci,v)
        _drow(ws2,r,10,alt=(ri%2==0))
    _widths(ws2,[20,16,16,14,14,12,14,14,12,10])

    # Sheet 3 — Alerts
    ws3=wb.create_sheet("Alert Produktivitas"); ws3.sheet_properties.tabColor="d64045"
    avg_ph=float(df["productivity_ton_per_ha"].mean()); thr=avg_ph*.75
    _title(ws3,"ALERT PRODUKTIVITAS — PT LONDON SUMATRA INDONESIA",7,"d64045")
    _sub(ws3,f"Threshold:<{thr:.3f} t/ha | Fleet avg:{avg_ph:.3f} | {gen}",7)
    for i,h in enumerate(["Tanggal","Estate","Produksi","Luas","Prod/ha","Fleet Avg","Defisit %"],1): ws3.cell(4,i,h)
    _hdr(ws3,4,7,"d64045")
    if len(alerts_df)>0:
        for ri,(_,row) in enumerate(alerts_df.iterrows()):
            r=ri+5; pv=float(row["productivity_ton_per_ha"]); deficit=round((avg_ph-pv)/avg_ph*100,1)
            try: ds=row["date"].strftime("%b %Y")
            except: ds=str(row["date"])
            for ci,v in enumerate([ds,str(row["estate"]),round(float(row["production_tons"]),2),
                                    round(float(row["plantation_area_ha"]),2),round(pv,4),round(avg_ph,4),deficit],1): ws3.cell(r,ci,v)
            _drow(ws3,r,7,alt=(ri%2==0))
            if deficit>40:
                for c in range(1,8): ws3.cell(r,c).fill=_hfill("FEE2E2")
    _widths(ws3,[14,18,14,12,16,16,12])

    # Sheet 4 — Forecast
    ws4=wb.create_sheet("Forecast 3 Bulan"); ws4.sheet_properties.tabColor="c9a84c"
    _title(ws4,"FORECAST 3 BULAN — PT LONDON SUMATRA INDONESIA",8,"c9a84c"); _sub(ws4,f"Model:{best} | {gen}",8)
    for i,h in enumerate(["Estate","Bulan+1 (ton)","Bulan+2 (ton)","Bulan+3 (ton)","Aktual","Tren M1%","Tren M3%"],1): ws4.cell(4,i,h)
    _hdr(ws4,4,8,"c9a84c")
    for ri,row in enumerate(forecast_df.itertuples()):
        r=ri+5
        for ci,v in enumerate([row.estate,round(row.m1,2),round(row.m2,2),round(row.m3,2),
                                round(row.last_actual,2),f"{row.chg_m1:+.1f}%",f"{row.chg_m3:+.1f}%"],1): ws4.cell(r,ci,v)
        _drow(ws4,r,8,alt=(ri%2==0))
    _widths(ws4,[20,16,16,16,18,14,14])

    # Sheet 5 — Model
    ws5=wb.create_sheet("Hasil Model ML"); ws5.sheet_properties.tabColor="457b9d"
    _title(ws5,"MODEL MACHINE LEARNING — PT LONDON SUMATRA INDONESIA",6); _sub(ws5,f"Model terbaik: {best} | {gen}",6)
    for i,h in enumerate(["Model","R²","MAE","RMSE","CV R²","Peringkat"],1): ws5.cell(4,i,h)
    _hdr(ws5,4,6,"0a1628")
    for ri,m in enumerate(model_results):
        r=ri+5
        for ci,v in enumerate([m["model"],m["r2"],m["mae"],m["rmse"],m["cv_r2"],ri+1],1): ws5.cell(r,ci,v)
        _drow(ws5,r,6,alt=(ri%2==0))
        if ri==0:
            for c in range(1,7): ws5.cell(r,c).fill=_hfill("D1FAE5"); ws5.cell(r,c).font=Font(bold=True,name="Calibri",color="065f46")
    _widths(ws5,[26,12,14,14,12,12])

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CORE PIPELINE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def process_dataset(raw: pd.DataFrame) -> dict:
    global _last
    dq = compute_dq(raw)

    df = raw.copy()
    for col in df.columns:
        if df[col].isnull().sum():
            df[col].fillna(df[col].median() if pd.api.types.is_numeric_dtype(df[col]) else df[col].mode()[0], inplace=True)
    df=df.drop_duplicates().reset_index(drop=True)
    df["date"]=pd.to_datetime(df["date"]); df=df.sort_values("date").reset_index(drop=True)
    df["year"]=df["date"].dt.year; df["month"]=df["date"].dt.month
    df["month_name"]=df["date"].dt.strftime("%b"); df["quarter"]=df["date"].dt.quarter
    df["productivity_ton_per_ha"]=(df["production_tons"]/df["plantation_area_ha"]).round(4)
    df["production_per_worker"]=(df["production_tons"]/df["workers"]).round(4)
    df["fertilizer_per_ha"]=(df["fertilizer_kg"]/df["plantation_area_ha"]).round(4)
    le=LabelEncoder(); df["estate_encoded"]=le.fit_transform(df["estate"])

    estates=sorted(df["estate"].unique().tolist())
    total_prod=float(df["production_tons"].sum()); avg_prod_ha=float(df["productivity_ton_per_ha"].mean())
    best_estate=str(df.groupby("estate")["production_tons"].sum().idxmax())
    peak_m=int(df.groupby("month")["production_tons"].mean().idxmax())
    date_range=df["date"].min().strftime("%b %Y")+" – "+df["date"].max().strftime("%b %Y")
    gen=datetime.now().strftime("%d %B %Y, %H:%M")

    kpis=dict(total_production_tons=round(total_prod,1),avg_productivity_t_ha=round(avg_prod_ha,4),
              best_estate=best_estate,peak_month=ML[peak_m-1],total_records=int(len(df)),
              date_range=date_range,estates=estates,num_estates=len(estates),generated_at=gen)
    alert_data=compute_alerts(df)

    FEAT=["plantation_area_ha","rainfall_mm","workers","fertilizer_kg","month","quarter","estate_encoded"]
    X,y=df[FEAT],df["production_tons"]
    Xtr,Xte,ytr,yte=train_test_split(X,y,test_size=.2,random_state=42)
    mdict={"Linear Regression":LinearRegression(),
           "Random Forest":RandomForestRegressor(n_estimators=200,random_state=42,n_jobs=-1),
           "Gradient Boosting":GradientBoostingRegressor(n_estimators=200,random_state=42)}
    results,trained={},{}
    for nm,mdl in mdict.items():
        mdl.fit(Xtr,ytr); yp=mdl.predict(Xte)
        results[nm]=dict(model=nm,mae=round(float(mean_absolute_error(yte,yp)),3),
                         rmse=round(float(np.sqrt(mean_squared_error(yte,yp))),3),
                         r2=round(float(r2_score(yte,yp)),4),
                         cv_r2=round(float(cross_val_score(mdl,X,y,cv=5,scoring="r2").mean()),4))
        trained[nm]=(mdl,yp)
    res_sorted=sorted(results.values(),key=lambda x:x["r2"],reverse=True)
    best_name=res_sorted[0]["model"]; best_mdl,best_pred=trained[best_name]; mae_val=res_sorted[0]["mae"]
    fi_series=(pd.Series(best_mdl.feature_importances_,index=FEAT) if hasattr(best_mdl,"feature_importances_")
               else pd.Series(np.abs(best_mdl.coef_)/np.abs(best_mdl.coef_).sum(),index=FEAT) if hasattr(best_mdl,"coef_") else None)
    fi_out={k:float(v) for k,v in fi_series.sort_values(ascending=False).items()} if fi_series is not None else {}

    fc_rows,chart_3m=compute_forecast_3m(df,best_mdl,le,FEAT,mae_val)
    fc_df=pd.DataFrame(fc_rows)
    thr=avg_prod_ha*.75; alerts_df=df[df["productivity_ton_per_ha"]<thr].sort_values("productivity_ton_per_ha").reset_index(drop=True)
    estate_stats={e:{"avg_area":float(df[df["estate"]==e]["plantation_area_ha"].mean()),
                     "avg_rainfall":float(df[df["estate"]==e]["rainfall_mm"].mean()),
                     "avg_workers":float(df[df["estate"]==e]["workers"].mean()),
                     "avg_fertilizer":float(df[df["estate"]==e]["fertilizer_kg"].mean())} for e in estates}

    # ── Build chat context for Chat feature ──────────────────────────────────
    chat_ctx = (
        f"Dataset: {date_range} | {len(estates)} estate: {', '.join(estates)}\n"
        f"Total produksi: {total_prod:,.1f} ton | Avg produktivitas: {avg_prod_ha:.4f} t/ha\n"
        f"Estate terbaik: {best_estate} | Bulan puncak: {ML[peak_m-1]}\n"
        f"Model terbaik: {best_name} (R²={res_sorted[0]['r2']:.4f}, MAE={mae_val:.3f} ton)\n"
        f"Top 3 faktor penting: {', '.join(list(fi_out.keys())[:3])}\n"
        f"Alert kritis: {[a['estate'] for a in alert_data if a['level']=='crit']}\n"
        f"Forecast bulan depan (total): {sum(r['m1'] for r in fc_rows):,.1f} ton\n"
        f"Data quality score: {dq['score']}/100\n\n"
        f"Ringkasan per estate:\n"
        +df.groupby("estate").agg(total=("production_tons","sum"),avg_ph=("productivity_ton_per_ha","mean")).round(3).to_string()
    )

    # ── CHARTS ───────────────────────────────────────────────────────────────
    monthly=df.groupby("date")["production_tons"].sum().reset_index()
    fig,ax=plt.subplots(figsize=(18,5.5))
    ax.fill_between(monthly["date"],monthly["production_tons"],alpha=.1,color=C["green"])
    ax.plot(monthly["date"],monthly["production_tons"],color=C["green"],lw=2,marker="o",ms=3.5,zorder=4,label="Total Bulanan")
    ax.plot(monthly["date"],monthly["production_tons"].rolling(3,min_periods=1).mean(),color=C["gold"],lw=2.5,ls="--",zorder=5,label="Rolling Avg 3 Bulan")
    ax.set_title("Tren Produksi Bulanan & Rata-rata Bergulir",fontsize=13,fontweight="bold",color=C["dark"])
    ax.set_xlabel("Tanggal"); ax.set_ylabel("Produksi (ton)"); ax.legend(fontsize=9)
    ax.tick_params(axis="x",rotation=25); fig.tight_layout(pad=1.5)
    c_trend=fig_b64(fig,dpi=140)

    mavg=df.groupby("month")["production_tons"].mean()
    fig,ax=plt.subplots(figsize=(10,5))
    bars=ax.bar(mavg.index,mavg.values,width=.65,edgecolor="white",
                color=[C["red"] if v==mavg.min() else (C["lime"] if v==mavg.max() else C["teal"]) for v in mavg.values],zorder=3)
    ax.set_xticks(range(1,13)); ax.set_xticklabels(ML)
    ax.axhline(mavg.mean(),color=C["gold"],ls="--",lw=1.5,label=f"Rata-rata ({mavg.mean():.1f}t)",zorder=4)
    ax.set_title("Profil Musiman — Avg Produksi per Bulan",fontsize=13,fontweight="bold",color=C["dark"])
    ax.set_ylabel("Produksi (ton)"); ax.legend(fontsize=9)
    _bar_labels(ax,bars)
    fig.tight_layout(pad=1.5); c_seasonal=fig_b64(fig,dpi=140)

    piv=df.groupby(["year","estate"])["production_tons"].sum().unstack(fill_value=0)
    fig,ax=plt.subplots(figsize=(10,5))
    piv.plot(kind="bar",stacked=True,ax=ax,color=PAL[:len(piv.columns)],edgecolor="white",width=.6)
    totals=piv.sum(axis=1)
    for i,(idx,tot) in enumerate(totals.items()):
        ax.text(i,tot+totals.max()*.012,f"{tot:,.0f}",ha="center",va="bottom",fontsize=8.5,fontweight="700",color=C["dark"])
    ax.set_title("Produksi Tahunan per Estate",fontsize=13,fontweight="bold",color=C["dark"])
    ax.set_xlabel("Tahun"); ax.set_ylabel("Total Produksi (ton)"); ax.tick_params(axis="x",rotation=0)
    ax.legend(title="Estate",bbox_to_anchor=(1.01,1),loc="upper left",fontsize=9)
    fig.tight_layout(pad=1.5); c_annual=fig_b64(fig,dpi=140)

    eo=df.groupby("estate")["production_tons"].median().sort_values(ascending=False).index
    fig,ax=plt.subplots(figsize=(10,5))
    bp=ax.boxplot([df[df["estate"]==e]["production_tons"].values for e in eo],patch_artist=True,labels=eo,widths=.55,
                  medianprops=dict(color=C["gold"],lw=2.5),whiskerprops=dict(color=C["gray"]),
                  capprops=dict(color=C["gray"]),flierprops=dict(marker="o",color=C["red"],markersize=4,alpha=.5))
    for i,(p,e) in enumerate(zip(bp["boxes"],eo)):
        p.set_facecolor(PAL[i%len(PAL)]); p.set_alpha(.7)
        med=float(df[df["estate"]==e]["production_tons"].median())
        ax.text(i+1,med,f"{med:.0f}",ha="center",va="bottom",fontsize=7.5,fontweight="700",color=C["dark"],zorder=6)
    ax.set_title("Distribusi Produksi per Estate",fontsize=13,fontweight="bold",color=C["dark"])
    ax.set_xlabel("Estate"); ax.set_ylabel("Produksi (ton)"); ax.tick_params(axis="x",rotation=20)
    fig.tight_layout(pad=1.5); c_boxplot=fig_b64(fig,dpi=140)

    pha=df.groupby("estate")["productivity_ton_per_ha"].mean().sort_values()
    fig,ax=plt.subplots(figsize=(10,5))
    b_ph=ax.barh(pha.index,pha.values,color=[C["red"] if v<float(pha.mean()) else C["green"] for v in pha.values],edgecolor="white",height=.55)
    ax.axvline(float(pha.mean()),color=C["gold"],ls="--",lw=2,label=f"Rata-rata ({pha.mean():.3f})",zorder=4)
    ax.set_title("Produktivitas per Hektar per Estate",fontsize=13,fontweight="bold",color=C["dark"])
    ax.set_xlabel("Ton / Ha"); ax.legend(fontsize=9)
    _barh_labels(ax,b_ph,fmt="{:.3f}")
    fig.tight_layout(pad=1.5); c_prodha=fig_b64(fig,dpi=140)

    ccols=["plantation_area_ha","rainfall_mm","workers","fertilizer_kg","productivity_ton_per_ha","production_tons"]
    corr=df[ccols].corr()
    fig,axes=plt.subplots(1,2,figsize=(16,6))
    sns.heatmap(corr,mask=np.triu(np.ones_like(corr,dtype=bool)),annot=True,fmt=".2f",
                cmap=sns.diverging_palette(10,150,s=80,as_cmap=True),ax=axes[0],
                linewidths=.5,cbar_kws={"shrink":.8},vmin=-1,vmax=1,annot_kws={"size":9,"weight":"600"})
    axes[0].set_title("Matriks Korelasi",fontsize=12,fontweight="bold",color=C["dark"])
    cp=corr["production_tons"].drop("production_tons").sort_values()
    axes[1].barh(cp.index,cp.values,color=[C["red"] if v<0 else C["green"] for v in cp.values],edgecolor="white",height=.6)
    axes[1].axvline(0,color=C["dark"],lw=1)
    axes[1].set_title("Korelasi dengan Produksi",fontsize=12,fontweight="bold",color=C["dark"])
    axes[1].set_xlabel("Koefisien Pearson")
    for i,v in enumerate(cp.values):
        axes[1].text(v+(.015 if v>=0 else -.015),i,f"{v:+.3f}",va="center",
                     ha="left" if v>=0 else "right",fontsize=9,fontweight="700",
                     color=C["green"] if v>=0 else C["red"])
    fig.tight_layout(pad=1.5); c_corr=fig_b64(fig,dpi=140)

    fig,axes=plt.subplots(2,2,figsize=(14,10))
    fig.suptitle("Driver Produksi — Scatter Analysis",fontsize=14,fontweight="bold",color=C["dark"],y=1.01)
    for ax,(x,xl,col) in zip(axes.flatten(),[("rainfall_mm","Curah Hujan (mm)",C["teal"]),
                                              ("fertilizer_kg","Pupuk (kg)",C["gold"]),
                                              ("workers","Jumlah Pekerja",C["navy"]),
                                              ("plantation_area_ha","Luas Lahan (ha)",C["green"])]):
        ax.scatter(df[x],df["production_tons"],alpha=.45,color=col,s=18,zorder=3,edgecolors="white",linewidths=.3)
        z=np.polyfit(df[x],df["production_tons"],1); xr=np.linspace(float(df[x].min()),float(df[x].max()),200)
        ax.plot(xr,np.poly1d(z)(xr),color=C["red"],lw=2,ls="--",label="Tren",zorder=4)
        r=float(df[[x,"production_tons"]].corr().iloc[0,1])
        ax.set_xlabel(xl); ax.set_ylabel("Produksi (ton)")
        ax.set_title(f"{xl}  (r={r:+.3f})",fontsize=10,fontweight="bold"); ax.legend(fontsize=8)
    fig.tight_layout(pad=1.5); c_scatter=fig_b64(fig,dpi=140)

    fig,axes=plt.subplots(1,3,figsize=(18,5))
    fig.suptitle(f"Evaluasi Model — {best_name}",fontsize=13,fontweight="bold",color=C["dark"])
    mn2=min(float(yte.min()),float(best_pred.min())); mx2=max(float(yte.max()),float(best_pred.max()))
    axes[0].scatter(yte,best_pred,alpha=.55,color=C["green"],s=22,zorder=3,edgecolors="white",linewidths=.3)
    axes[0].plot([mn2,mx2],[mn2,mx2],color=C["red"],lw=2,ls="--",label="Ideal")
    axes[0].text(.05,.90,f"R² = {res_sorted[0]['r2']:.4f}",transform=axes[0].transAxes,fontsize=12,color=C["red"],fontweight="bold")
    axes[0].set_xlabel("Aktual (ton)"); axes[0].set_ylabel("Prediksi (ton)"); axes[0].set_title("Aktual vs Prediksi",fontweight="bold"); axes[0].legend()
    resid=yte.values-best_pred
    axes[1].scatter(best_pred,resid,alpha=.5,color=C["gold"],s=22,zorder=3,edgecolors="white",linewidths=.3)
    axes[1].axhline(0,color=C["red"],lw=2,ls="--"); axes[1].set_xlabel("Prediksi"); axes[1].set_ylabel("Residual"); axes[1].set_title("Plot Residual",fontweight="bold")
    axes[2].hist(resid,bins=28,color=C["teal"],edgecolor="white",alpha=.85); axes[2].axvline(0,color=C["red"],lw=2,ls="--")
    axes[2].set_xlabel("Residual"); axes[2].set_ylabel("Frekuensi"); axes[2].set_title("Distribusi Residual",fontweight="bold")
    fig.tight_layout(pad=1.5); c_model_eval=fig_b64(fig,dpi=140)

    c_fi=""
    if fi_series is not None:
        fi_s=fi_series.sort_values(ascending=True)
        fig,ax=plt.subplots(figsize=(12,6))
        b_fi=ax.barh(fi_s.index,fi_s.values,color=[C["green"] if v>=float(fi_s.median()) else C["teal"] for v in fi_s.values],edgecolor="white",height=.6)
        ax.set_title(f"Feature Importance — {best_name}",fontsize=13,fontweight="bold",color=C["dark"])
        ax.set_xlabel("Skor Kepentingan")
        _barh_labels(ax,b_fi,fmt="{:.4f}",fs=9.5)
        ax.set_xlim(0,fi_s.max()*1.18); fig.tight_layout(pad=1.5); c_fi=fig_b64(fig,dpi=140)

    # [FIX-2] Parallel AI insights
    estate_str=df.groupby("estate").agg(total=("production_tons","sum"),avg_ph=("productivity_ton_per_ha","mean")).round(3).to_string()
    corr_str=corr["production_tons"].drop("production_tons").round(3).to_string()
    prompts={
        "trend":    f"Tren produksi Lonsum: total={total_prod:,.1f}t | {date_range} | estate={','.join(estates)}\nAnalisis tren utama. 1 rekomendasi.",
        "seasonal": f"Pola musiman:\n{df.groupby('month_name')['production_tons'].mean().round(1).to_string()}\nBulan puncak:{ML[peak_m-1]}. Implikasi. 1 rekomendasi.",
        "annual":   f"Produksi tahunan per estate:\n{estate_str}\nBandingkan performa. 1 rekomendasi.",
        "boxplot":  f"Distribusi produksi:\n{estate_str}\nEstate paling stabil? 1 rekomendasi.",
        "prodha":   f"Produktivitas/ha:\n{df.groupby('estate')['productivity_ton_per_ha'].mean().round(4).to_string()}\nFleet avg:{avg_prod_ha:.4f}. 1 rekomendasi.",
        "correlation": f"Korelasi:\n{corr_str}\nFaktor terpenting? 1 rekomendasi.",
        "scatter":  f"Driver: hujan, pupuk, pekerja, lahan. {date_range}. Driver paling actionable? 1 rekomendasi.",
        "model":    f"Model {best_name}: R²={res_sorted[0]['r2']:.4f}, MAE={mae_val:.3f}t. Jelaskan. 1 rekomendasi.",
        "feature_importance": "\n".join([f"  {k}:{v:.4f}" for k,v in list(fi_out.items())[:5]])+"\nFaktor paling penting? 1 rekomendasi.",
        "forecast": f"Forecast:\n{pd.DataFrame(fc_rows)[['estate','m1','m2','m3']].to_string()}\nModel={best_name}, MAE={mae_val:.3f}t. 1 rekomendasi.",
    }
    ai_insights = ask_llm_parallel(prompts)

    charts={
        "trend":c_trend,"seasonal":c_seasonal,"annual":c_annual,
        "boxplot":c_boxplot,"prodha":c_prodha,"corr":c_corr,
        "scatter":c_scatter,"model_eval":c_model_eval,"feature_imp":c_fi,
        "forecast":chart_3m,"dq":dq.get("chart",""),
    }
    # Auto comparative kalau ada 2+ tahun
    comp_result=None
    years=sorted(df["year"].unique())
    if len(years)>=2:
        dfa=df[df["year"]==years[-2]].copy()
        dfb=df[df["year"]==years[-1]].copy()
        try: comp_result=compute_comparative(dfa,dfb,str(years[-2]),str(years[-1]))
        except: pass

    result={"kpis":kpis,"model_results":res_sorted,"best_model":best_name,
            "feature_importance":fi_out,"forecast_3m":fc_rows,"alert_data":alert_data,
            "data_quality":dq,"charts":charts,"ai_insights":ai_insights,
            "generated_at":gen,"estate_stats":estate_stats,
            "comparative":comp_result}
    _last={**result,"_df":df,"_le":le,"_best_mdl":best_mdl,"_FEAT":FEAT,"_mae":mae_val,
          "_fc_df":fc_df,"_alerts_df":alerts_df,"_chat_ctx":chat_ctx}
    return result


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# HTML — FRONTEND  [FIX-7] CSS consolidated, dark topbar, chat widget
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
HTML_PAGE = r"""<!DOCTYPE html>
<html lang="id">
<head>
<meta charset="UTF-8"/><meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>Lonsum LEAP v5.0 — Plantation Intelligence</title>
<link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700;800&family=Fraunces:ital,opsz,wght@0,9..144,300;0,9..144,600;1,9..144,300&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet"/>
<style>
/* ── Reset & Variables ── */
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{
  --dk:#0a1628;--nv:#0d2137;--gr:#1a6b3c;--tl:#0e7c6e;
  --go:#c9a84c;--li:#3dba6f;--rd:#d64045;--or:#e07b39;
  --lt:#e8f5ee;--bg:#f0f4f8;--bd:#dde3ec;--tx:#18243a;--mu:#637289;
  --sw:260px;
  --xs:0 1px 3px rgba(0,0,0,.06);--sm:0 2px 8px rgba(13,33,55,.08);
  --md:0 4px 20px rgba(13,33,55,.11);--lg:0 12px 48px rgba(13,33,55,.16);
  --r:10px;--rl:14px;--rx:18px;
}
body.dark{
  --bg:#0d1b2e;--bd:#1e3048;--tx:#e2eaf4;--mu:#7a92b0;
  --lt:#0f2a1e;--nv:#060f1c;
}
body.dark .kpi,
body.dark .ac,
body.dark .ml-tbl,
body.dark .fc-tbl,
body.dark .sim-card,
body.dark .comp-c,
body.dark .up-card,
body.dark .meta-r{background:#0d2137;border-color:#1e3048;color:#e2eaf4}
body.dark .ml-tbl tbody tr:hover,
body.dark .fc-tbl tbody tr:hover{background:#0f2a40}
body.dark .sf input,
body.dark .sf select{background:#0a1628;border-color:#1e3048;color:#e2eaf4}
body.dark .ac-ins{background:linear-gradient(135deg,#0f2a1e,#0a1f17);border-color:rgba(26,107,60,.2)}
body.dark .ac-ins p{color:#a7c4b0}
body.dark .ch-inp{background:#0d2137}
body.dark .ch-inp input{background:#0a1628;border-color:#1e3048;color:#e2eaf4}
body.dark .msg.ai .msg-bub{background:#0d2137;border-color:#1e3048;color:#e2eaf4}
body.dark #chat-win{background:#0a1628;border-color:#1e3048}
body.dark .ch-qs{background:#060f1c}
body.dark .up-card p{color:#7a92b0}
body.dark .col-hint code{background:#0f2a1e;color:#3dba6f}
body.dark .ph h2,
body.dark .ph p,
body.dark .ph .ph-l,
body.dark .sep .sl,
body.dark .ac-hdr h3,
body.dark .ac-hdr p,
body.dark .kpi-v,
body.dark .kpi-lb,
body.dark .kpi-s,
body.dark .dq-v,
body.dark .dq-l,
body.dark .mk-v,
body.dark .mk-l,
body.dark .sr-l h4,
body.dark .sr-u,
body.dark .al-b h4,
body.dark .al-b p,
body.dark .dl-l h3,
body.dark .dl-l p,
body.dark .sb-lbl,
body.dark .sf label,
body.dark .fc-tbl tbody td,
body.dark .ml-tbl tbody td,
body.dark .tb-bc,
body.dark .meta-r,
body.dark .meta-r strong,
body.dark .up-card h2,
body.dark .sim-body,
body.dark .ac-tag{color:#e2eaf4}

body.dark .ml-tbl thead th,
body.dark .fc-tbl thead th{color:rgba(255,255,255,.8)}

body.dark .sb-btn{color:rgba(255,255,255,.6)}
body.dark .sb-btn:hover{color:#fff}
body.dark .sb-btn.on{color:#fff}

body.dark .sep .ln{background:#1e3048}
body.dark .dq-tr{background:#1e3048}
body.dark .up-tab{background:#0d2137;border-color:#1e3048;color:#7a92b0}
body.dark .up-tab.on{background:#1a6b3c;color:#fff;border-color:#1a6b3c}
body.dark .chip span{color:rgba(255,255,255,.4)}
body.dark .btn-dlc{background:#0d2137;border-color:#1e3048;color:#7a92b0}
body.dark .btn-dlc:hover{background:#0f2a1e;color:#3dba6f}
body.dark .cq{background:#0d2137;border-color:#1e3048;color:#3dba6f}
body.dark .fmt-in{background:rgba(201,168,76,.06);border-color:rgba(201,168,76,.15)}
body.dark .info-box{background:rgba(255,255,255,.02);border-color:rgba(255,255,255,.06)}
body.dark .st{color:rgba(255,255,255,.85)}
body.dark .sd{color:rgba(255,255,255,.35)}
body.dark #pane-comp > div{background:#0d2137;border-color:#1e3048}
body.dark #pane-comp > div > div[style*="font-family"]{color:#e2eaf4}
body.dark #pane-comp p{color:#7a92b0}
body.dark #pane-comp > div > div[style*="font-size:.83rem"]{color:#7a92b0}
body.dark #pane-comp > div > div[style*="text-align:center"]{color:#7a92b0}
body.dark .cs{background:#0a1628;border-color:#1e3048}
body.dark .cs:hover,
body.dark .cs.filled{background:#0f2a1e;border-color:#1a6b3c}
body.dark .cs-l{color:#e2eaf4}
body.dark .cs-f{color:#3dba6f}
body.dark .cs-i{color:#e2eaf4}
body.dark #pane-comp [style*="color:var(--mu)"]{color:#7a92b0!important}
body.dark #pane-comp [style*="font-size:.75rem"]{color:#7a92b0!important}
body.dark #pane-comp [style*="font-size:.83rem"]{color:#7a92b0!important}
body.dark #pane-comp strong{color:#e2eaf4!important}
body.dark #ld-msg{color:#e2eaf4}
body.dark .ls{color:#7a92b0}
body.dark .ls.on{color:#3dba6f;background:rgba(26,107,60,.15)}
body.dark .ls.done{color:#3a5a47}
body.dark .pb{background:#1e3048}

/* Alert cards */
body.dark .al.crit{background:rgba(214,64,69,.15);border-color:rgba(214,64,69,.3);color:#fca5a5}
body.dark .al.crit .al-b h4,
body.dark .al.crit .al-b p{color:#fca5a5}
body.dark .al.warn{background:rgba(201,168,76,.15);border-color:rgba(201,168,76,.3);color:#fde68a}
body.dark .al.warn .al-b h4,
body.dark .al.warn .al-b p{color:#fde68a}
body.dark .al.ok{background:rgba(26,107,60,.15);border-color:rgba(26,107,60,.3);color:#86efac}
body.dark .al.ok .al-b h4,
body.dark .al.ok .al-b p{color:#86efac}

/* ML table best row */
body.dark .ml-tbl tbody tr[style*="background:#f0faf5"]{background:#0f2a1e!important}
body.dark .ml-tbl tbody tr[style*="background:#f0faf5"] td{color:#e2eaf4!important}
body.dark .ml-tbl tbody tr[style*="background:#f0faf5"] strong{color:#3dba6f!important}

/* Comparative cards */
body.dark .comp-c > div > div[style*="background:#f0fdf4"]{background:rgba(26,107,60,.2)!important}
body.dark .comp-c > div > div[style*="background:#fff1f0"]{background:rgba(214,64,69,.15)!important}
body.dark .comp-c > div > div[style*="background:var(--bg)"]{background:#0a1628!important}
body.dark .comp-c > div > div > div[style*="font-family"]{color:#e2eaf4!important}
body.dark .comp-c > div > div > div > div[style*="font-size:.7rem"]{color:#7a92b0!important}
body.dark .comp-c > div > div[style*="color:#166534"] > div{color:#86efac!important}
body.dark .comp-c > div > div[style*="color:#991b1b"] > div{color:#fca5a5!important}
/* Data Quality cards */
body.dark .dq-c{background:#0d2137;border-color:#1e3048}
body.dark .dq-v{color:#e2eaf4}
body.dark .dq-l{color:#7a92b0}
body.dark .dq-row .dq-c .dq-v[style*="color:#1a6b3c"]{color:#3dba6f!important}
body.dark .dq-row .dq-c .dq-v[style*="color:#c9a84c"]{color:#f0d080!important}
body.dark .dq-row .dq-c .dq-v[style*="color:#d64045"]{color:#fca5a5!important}
body.dark .dq-tr{background:#1e3048}

body.dark .ls .ld{background:transparent}
body.dark .ls.on .ld{background:var(--gr)}
body.dark .ls.done .ld{background:#3dba6f;color:#fff}
html{scroll-behavior:smooth}
body{font-family:'Plus Jakarta Sans',sans-serif;background:var(--bg);color:var(--tx);min-height:100vh;display:flex;-webkit-font-smoothing:antialiased}

/* ── LANDING ── */
#lp{position:fixed;inset:0;z-index:9999;overflow-y:auto;background:var(--dk)}
#lp.out{animation:lpOut .6s ease forwards;pointer-events:none}
@keyframes lpOut{to{opacity:0;transform:scale(1.04)}}
.lp-mesh{position:fixed;inset:-40%;pointer-events:none;z-index:0;
  background:radial-gradient(ellipse 65% 55% at 18% 28%,rgba(26,107,60,.38) 0%,transparent 60%),
  radial-gradient(ellipse 55% 65% at 82% 72%,rgba(14,124,110,.28) 0%,transparent 60%),
  radial-gradient(ellipse 45% 40% at 65% 18%,rgba(201,168,76,.14) 0%,transparent 55%);
  animation:mesh 20s ease-in-out infinite alternate}
@keyframes mesh{to{transform:translate(3%,2.5%)}}
.lp-grid{position:fixed;inset:0;pointer-events:none;z-index:0;
  background-image:linear-gradient(rgba(255,255,255,.025) 1px,transparent 1px),
    linear-gradient(90deg,rgba(255,255,255,.025) 1px,transparent 1px);background-size:52px 52px}
.orb{position:fixed;border-radius:50%;filter:blur(70px);pointer-events:none;z-index:0}
.o1{width:420px;height:420px;background:rgba(26,107,60,.22);top:-100px;left:-80px}
.o2{width:320px;height:320px;background:rgba(14,124,110,.2);bottom:5%;right:-60px}
.o3{width:220px;height:220px;background:rgba(201,168,76,.1);top:38%;left:58%}
.lp-wrap{position:relative;z-index:10;min-height:100vh;display:flex;flex-direction:column}

/* landing nav */
.lp-nav{display:flex;align-items:center;justify-content:space-between;padding:1.4rem 3rem;border-bottom:1px solid rgba(255,255,255,.06)}
.lp-brand{display:flex;align-items:center;gap:13px}
.lp-logo{width:44px;height:44px;border-radius:12px;background:linear-gradient(135deg,#1a6b3c,#0e7c6e);display:flex;align-items:center;justify-content:center;box-shadow:0 4px 20px rgba(26,107,60,.5);flex-shrink:0}
.lp-brand h1{font-family:'Fraunces',serif;font-size:1rem;font-weight:600;color:#fff}
.lp-brand p{font-size:.58rem;color:var(--go);font-weight:700;text-transform:uppercase;letter-spacing:.16em;margin-top:1px}
.pill{font-size:.63rem;font-weight:700;text-transform:uppercase;letter-spacing:.08em;padding:5px 13px;border-radius:20px}
.p-n{color:rgba(255,255,255,.4);border:1px solid rgba(255,255,255,.1)}
.p-g{color:var(--go);border:1px solid rgba(201,168,76,.3);background:rgba(201,168,76,.07)}

/* landing hero */
.lp-main{flex:1;display:flex;flex-direction:column;align-items:center;padding:4rem 2rem 3rem}
@keyframes up{from{opacity:0;transform:translateY(20px)}to{opacity:1;transform:none}}
.eyebrow{display:inline-flex;align-items:center;gap:9px;background:rgba(26,107,60,.18);border:1px solid rgba(26,107,60,.38);color:var(--li);font-size:.7rem;font-weight:700;letter-spacing:.07em;text-transform:uppercase;padding:6px 18px;border-radius:20px;margin-bottom:2rem;animation:up .7s .05s both}
.dot-pulse{width:7px;height:7px;border-radius:50%;background:var(--li);animation:blink 2s infinite;flex-shrink:0}
@keyframes blink{0%,100%{opacity:1}50%{opacity:.3}}
.lp-h1{font-family:'Fraunces',serif;font-size:clamp(2.8rem,6vw,5rem);font-weight:600;color:#fff;letter-spacing:-.04em;line-height:1.07;text-align:center;margin-bottom:.8rem;animation:up .7s .12s both}
.acc{background:linear-gradient(130deg,var(--go) 0%,var(--li) 65%);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text}
.it{font-style:italic;font-weight:300;color:rgba(255,255,255,.65)}
.lp-tag{font-size:clamp(.88rem,1.8vw,1.05rem);color:rgba(255,255,255,.48);max-width:560px;text-align:center;line-height:1.82;margin-bottom:2.8rem;animation:up .7s .2s both}
.lp-tag strong{color:rgba(255,255,255,.8);font-weight:600}
.btn-start{background:linear-gradient(135deg,#1a6b3c,#0e7c6e);color:#fff;padding:16px 48px;border-radius:12px;font-size:.98rem;font-weight:700;border:none;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;box-shadow:0 8px 32px rgba(26,107,60,.5);transition:all .25s;display:inline-flex;align-items:center;gap:10px;animation:up .7s .28s both;margin-bottom:3.5rem}
.btn-start:hover{transform:translateY(-3px);box-shadow:0 14px 44px rgba(26,107,60,.62)}
.arr{transition:transform .2s}.btn-start:hover .arr{transform:translateX(5px)}
.feat-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:1rem;width:100%;max-width:960px;margin-bottom:2.5rem;animation:up .7s .35s both}
@media(max-width:760px){.feat-grid{grid-template-columns:repeat(2,1fr)}}
.feat{background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.08);border-radius:14px;padding:1.2rem 1.1rem;transition:all .22s}
.feat:hover{background:rgba(255,255,255,.07);transform:translateY(-3px)}
.feat-i{font-size:1.4rem;margin-bottom:.6rem}
.feat-t{font-size:.8rem;font-weight:700;color:#fff;margin-bottom:.25rem}
.feat-d{font-size:.7rem;color:rgba(255,255,255,.38);line-height:1.65}
.new-badge{display:inline-block;font-size:.55rem;font-weight:800;text-transform:uppercase;letter-spacing:.1em;background:rgba(61,186,111,.2);color:var(--li);border:1px solid rgba(61,186,111,.35);padding:2px 7px;border-radius:20px;margin-left:5px;vertical-align:middle}

/* steps & format */
.info-box{width:100%;max-width:860px;margin-bottom:1.8rem;background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.07);border-radius:16px;padding:1.8rem 2rem;animation:up .7s .41s both}
.sec-lbl{font-size:.62rem;font-weight:700;text-transform:uppercase;letter-spacing:.15em;color:var(--go);margin-bottom:1.2rem;display:flex;align-items:center;gap:8px}
.sec-lbl::before{content:'';display:inline-block;width:16px;height:2px;background:var(--go);border-radius:1px}
.steps{display:grid;grid-template-columns:repeat(4,1fr);gap:1.2rem}
.sn{width:30px;height:30px;border-radius:9px;margin-bottom:.55rem;background:linear-gradient(135deg,var(--gr),var(--tl));display:inline-flex;align-items:center;justify-content:center;font-size:.72rem;font-weight:800;color:#fff}
.st{font-size:.78rem;font-weight:700;color:rgba(255,255,255,.85);margin-bottom:.22rem}
.sd{font-size:.68rem;color:rgba(255,255,255,.37);line-height:1.6}
.fmt-box{width:100%;max-width:860px;margin-bottom:1.8rem;animation:up .7s .46s both}
.fmt-in{background:rgba(201,168,76,.06);border:1px solid rgba(201,168,76,.2);border-radius:14px;padding:1.1rem 1.7rem}
.chips{display:flex;flex-wrap:wrap;gap:.45rem;margin-top:.65rem}
.chip{display:flex;align-items:center;gap:7px;background:rgba(255,255,255,.05);border:1px solid rgba(255,255,255,.1);border-radius:8px;padding:.38rem .85rem}
.chip code{font-family:'JetBrains Mono',monospace;font-size:.67rem;color:var(--li)}
.chip span{font-size:.63rem;color:rgba(255,255,255,.28)}
.lp-foot{border-top:1px solid rgba(255,255,255,.06);padding:1rem 3rem;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:.5rem}
.lp-foot p{font-size:.67rem;color:rgba(255,255,255,.22)}
.stags{display:flex;gap:.4rem;flex-wrap:wrap}
.stag{font-family:'JetBrains Mono',monospace;font-size:.58rem;color:rgba(255,255,255,.18);border:1px solid rgba(255,255,255,.07);padding:3px 8px;border-radius:20px}

/* ── SIDEBAR ── */
#sb{width:var(--sw);min-height:100vh;background:var(--dk);display:flex;flex-direction:column;position:fixed;top:0;left:0;bottom:0;z-index:300;border-right:1px solid rgba(255,255,255,.06);overflow:hidden;transition:transform .3s;flex-shrink:0}
.sb-top{padding:1.5rem 1.4rem 1.2rem;border-bottom:1px solid rgba(255,255,255,.07)}
.sb-logo{display:flex;align-items:center;gap:12px;margin-bottom:1.1rem}
.sb-li{width:42px;height:42px;border-radius:10px;background:linear-gradient(135deg,var(--gr),var(--tl));display:flex;align-items:center;justify-content:center;box-shadow:0 4px 14px rgba(26,107,60,.45);flex-shrink:0}
.sb-brand h1{font-family:'Fraunces',serif;font-size:1.1rem;font-weight:600;color:#fff}
.sb-brand p{font-size:.62rem;color:var(--go);font-weight:600;text-transform:uppercase;letter-spacing:.13em;margin-top:2px}
.sb-mi{font-size:.7rem;color:rgba(255,255,255,.4);display:flex;align-items:center;gap:6px;margin-top:4px}
.sb-mi span:first-child{color:rgba(255,255,255,.2)}
.sb-nav{flex:1;padding:1.2rem .8rem;overflow-y:auto;scrollbar-width:none}
.sb-nav::-webkit-scrollbar{display:none}
.sb-sec{margin-bottom:1.4rem}
.sb-lbl{font-size:.6rem;font-weight:700;text-transform:uppercase;letter-spacing:.14em;color:rgba(255,255,255,.25);padding:.2rem .7rem .5rem;display:block}
.sb-btn{display:flex;align-items:center;gap:10px;width:100%;padding:.65rem .85rem;border-radius:var(--r);color:rgba(255,255,255,.55);font-size:.8rem;font-weight:500;cursor:pointer;border:none;background:transparent;text-align:left;transition:all .18s;font-family:'Plus Jakarta Sans',sans-serif}
.sb-btn:hover{background:rgba(255,255,255,.06);color:rgba(255,255,255,.9)}
.sb-btn.on{background:rgba(26,107,60,.25);color:#fff;border-left:3px solid var(--li)}
.sb-btn.on .si{color:var(--li)}
.si{font-size:1rem;width:20px;text-align:center;flex-shrink:0}
.sb-bx{margin-left:auto;font-size:.58rem;background:rgba(201,168,76,.2);color:var(--go);padding:2px 7px;border-radius:20px;font-weight:700}
.sb-bn{margin-left:auto;font-size:.55rem;background:rgba(61,186,111,.2);color:var(--li);padding:2px 7px;border-radius:20px;font-weight:800;text-transform:uppercase;letter-spacing:.05em}
.sb-bot{padding:1rem .8rem 1.4rem;border-top:1px solid rgba(255,255,255,.07)}
.sb-ver{font-size:.6rem;color:rgba(255,255,255,.2);text-align:center;margin-top:4px}

/* ── MAIN ── */
#main{margin-left:var(--sw);flex:1;min-height:100vh;display:flex;flex-direction:column;overflow-x:hidden}
#main.chat-open{margin-right:420px;transition:margin-right .25s ease}

/* [FIX-10] Dark topbar */
#tb{background:var(--dk);border-bottom:1px solid rgba(255,255,255,.08);height:62px;display:flex;align-items:center;justify-content:space-between;padding:0 2rem;position:sticky;top:0;z-index:100;box-shadow:0 2px 12px rgba(0,0,0,.25)}
.tb-l{display:flex;align-items:center;gap:1rem}
.tb-l h2{font-family:'Fraunces',serif;font-size:1.15rem;font-weight:600;color:#fff;letter-spacing:-.01em}
.tb-bc{font-size:.75rem;color:rgba(255,255,255,.35)}
.tb-r{display:flex;align-items:center;gap:.7rem}
.ai-badge{background:linear-gradient(135deg,var(--go),#e8c45a);color:var(--dk);font-size:.65rem;font-weight:700;padding:4px 12px;border-radius:20px;text-transform:uppercase;letter-spacing:.04em}
.btn-reset{background:rgba(255,255,255,.08);border:1px solid rgba(255,255,255,.15);color:rgba(255,255,255,.7);padding:6px 14px;border-radius:var(--r);font-size:.76rem;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;font-weight:600;transition:all .2s;display:none}
.btn-reset:hover{background:rgba(255,255,255,.15);color:#fff}
#cnt{flex:1;padding:2rem 2rem 6rem 4rem}

/* ── UPLOAD ── */
#up-sec{max-width:800px;margin:1rem auto 0}
.up-tabs{display:flex;gap:.5rem;margin-bottom:1.5rem}
.up-tab{padding:9px 20px;border-radius:var(--r);font-size:.8rem;font-weight:700;cursor:pointer;border:1px solid var(--bd);background:#fff;color:var(--mu);transition:all .2s;font-family:'Plus Jakarta Sans',sans-serif}
.up-tab.on{background:var(--dk);color:#fff;border-color:var(--dk)}
.up-pane{display:none}.up-pane.on{display:block}
.up-card{background:#fff;border:2px dashed var(--bd);border-radius:var(--rx);padding:3rem 2.5rem;text-align:center;cursor:pointer;transition:all .25s;box-shadow:var(--sm)}
.up-card:hover,.up-card.dv{border-color:var(--gr);border-style:solid;background:var(--lt);box-shadow:var(--md);transform:translateY(-3px)}
.up-icon{font-size:3rem;margin-bottom:1rem}
.up-card h2{font-family:'Fraunces',serif;font-size:1.5rem;font-weight:600;color:var(--dk);margin-bottom:.5rem}
.up-card p{color:var(--mu);font-size:.85rem;line-height:1.7}
.col-hint{display:flex;flex-wrap:wrap;justify-content:center;gap:6px;margin-top:1rem}
.col-hint code{font-family:'JetBrains Mono',monospace;font-size:.7rem;background:var(--lt);color:var(--gr);padding:4px 10px;border-radius:6px;border:1px solid rgba(26,107,60,.2)}
.btn-p{background:linear-gradient(135deg,var(--gr),var(--tl));color:#fff;padding:12px 32px;border-radius:var(--r);margin-top:1.5rem;font-weight:700;font-size:.88rem;cursor:pointer;border:none;font-family:'Plus Jakarta Sans',sans-serif;box-shadow:0 4px 18px rgba(26,107,60,.3);transition:all .2s;display:inline-flex;align-items:center;gap:8px}
.btn-p:hover{transform:translateY(-2px);box-shadow:0 6px 24px rgba(26,107,60,.4)}
.btn-p:disabled{opacity:.5;pointer-events:none}
#fi,#fia,#fib{display:none}
.err-b{background:#fff5f5;border:1px solid #fecaca;border-radius:var(--r);padding:12px 16px;margin-bottom:16px;color:var(--rd);font-size:.85rem}
.cp2{display:grid;grid-template-columns:1fr 1fr;gap:1rem;margin-bottom:1rem}
.cs{background:var(--bg);border:2px dashed var(--bd);border-radius:var(--rl);padding:1.2rem;text-align:center;cursor:pointer;transition:all .2s}
.cs:hover,.cs.filled{border-color:var(--gr);border-style:solid;background:var(--lt)}
.cs-i{font-size:1.8rem;margin-bottom:.4rem}
.cs-l{font-size:.75rem;font-weight:700;color:var(--dk);margin-bottom:.2rem}
.cs-f{font-size:.7rem;color:var(--gr);font-weight:600}

/* ── LOADING ── */
#ld{display:none;flex-direction:column;align-items:center;justify-content:center;min-height:calc(100vh - 62px);gap:1.5rem;padding:2rem}
.sp-w{position:relative;width:72px;height:72px}
.sp{width:72px;height:72px;border:3px solid var(--bd);border-top-color:var(--gr);border-radius:50%;animation:spin .8s linear infinite}
.spi{position:absolute;top:12px;left:12px;width:48px;height:48px;border:3px solid transparent;border-top-color:var(--go);border-radius:50%;animation:spin 1.2s linear infinite reverse}
@keyframes spin{to{transform:rotate(360deg)}}
#ld-msg{font-family:'Fraunces',serif;font-size:1.2rem;color:var(--dk);text-align:center}
.pb{width:320px;height:5px;background:var(--bd);border-radius:3px;overflow:hidden}
.pbf{height:100%;background:linear-gradient(90deg,var(--gr),var(--go));border-radius:3px;animation:prog 28s ease-out forwards}
@keyframes prog{0%{width:2%}50%{width:60%}90%{width:88%}100%{width:95%}}
.ls-wrap{display:flex;flex-direction:column;gap:4px;width:340px}
.ls{font-size:.78rem;color:var(--mu);padding:6px 12px;border-radius:8px;transition:all .3s;display:flex;align-items:center;justify-content:space-between;gap:10px}
.ls.on{color:var(--gr);background:var(--lt);font-weight:700}
.ls.done{color:#3dba6f;font-weight:600}
.ls{font-size:.78rem;color:var(--mu);padding:6px 12px;border-radius:8px;transition:all .3s;display:flex;align-items:center;justify-content:space-between;gap:10px}
.ls.on{color:var(--gr);background:var(--lt);font-weight:700}
.ls.done{color:#3dba6f;font-weight:600}
.ck{width:22px;height:22px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:.9rem;flex-shrink:0;opacity:0;transition:all .3s}
.ls.done .ck{background:#3dba6f;color:#fff;opacity:1}
.ld{width:6px;height:6px;border-radius:50%;background:currentColor;flex-shrink:0}

/* ── DASHBOARD ── */
#dash{display:none;animation:up .45s ease;padding-top:0;margin-top:0}
.ph{margin-bottom:1.8rem}
.ph .ph-l{font-size:.65rem;font-weight:700;color:var(--go);text-transform:uppercase;letter-spacing:.14em;margin-bottom:.3rem}
.ph h2{font-family:'Fraunces',serif;font-size:1.7rem;font-weight:600;color:var(--dk);line-height:1.2}
.ph p{color:var(--mu);font-size:.84rem;margin-top:.3rem}
.sep{margin:2.2rem 0 1.4rem;display:flex;align-items:center;gap:12px}
.sep .sl{font-size:.63rem;font-weight:700;color:var(--go);text-transform:uppercase;letter-spacing:.14em;white-space:nowrap}
.sep .ln{flex:1;height:1px;background:var(--bd)}

/* alerts */
.al{border-radius:var(--rl);padding:.9rem 1.2rem;margin-bottom:.6rem;display:flex;align-items:flex-start;gap:12px;border:1px solid}
.al.crit{background:#fff1f1;border-color:#fecaca;color:#991b1b}
.al.warn{background:#fffbeb;border-color:#fde68a;color:#92400e}
.al.ok{background:#f0fdf4;border-color:#bbf7d0;color:#166534}
.al-ico{font-size:1.2rem;flex-shrink:0}
.al-b h4{font-size:.82rem;font-weight:700;margin-bottom:2px}
.al-b p{font-size:.76rem;line-height:1.6}

/* DQ */
.dq-row{display:grid;grid-template-columns:repeat(4,1fr);gap:1rem;margin-bottom:1.8rem}
.dq-c{background:#fff;border-radius:var(--rl);padding:1.1rem;border:1px solid var(--bd);box-shadow:var(--xs);text-align:center}
.dq-v{font-family:'Fraunces',serif;font-size:2rem;font-weight:600;line-height:1}
.dq-l{font-size:.7rem;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:var(--mu);margin-top:.3rem}
.dq-tr{height:5px;background:var(--lt);border-radius:3px;margin-top:.5rem;overflow:hidden}
.dq-tf{height:100%;border-radius:3px}

/* KPIs */
.kpi-g{display:grid;grid-template-columns:repeat(4,1fr);gap:1rem;margin-bottom:1.8rem;min-width:0}
@media(max-width:1400px){.kpi-g{grid-template-columns:repeat(3,1fr)}}
@media(max-width:1100px){.kpi-g{grid-template-columns:repeat(2,1fr)}}
.kpi{background:#fff;border-radius:var(--rl);padding:1.4rem 1.3rem 1.2rem;box-shadow:var(--xs);border:1px solid var(--bd);position:relative;overflow:hidden;transition:all .22s;cursor:default}
.kpi:hover{box-shadow:var(--md);transform:translateY(-2px)}
.kpi-ac{position:absolute;top:0;left:0;right:0;height:4px;border-radius:var(--rl) var(--rl) 0 0}
.ac-g{background:linear-gradient(90deg,var(--gr),var(--tl))}
.ac-go{background:linear-gradient(90deg,var(--go),#e8c45a)}
.ac-t{background:linear-gradient(90deg,var(--tl),var(--li))}
.ac-r{background:linear-gradient(90deg,var(--or),var(--rd))}
.kpi-top{display:flex;align-items:flex-start;justify-content:space-between;margin-bottom:.7rem;margin-top:.2rem}
.kpi-iw{width:38px;height:38px;border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:1.1rem;background:var(--lt)}
.kpi-v{font-family:'Fraunces',serif;font-size:1.9rem;font-weight:600;color:var(--dk);line-height:1;margin-bottom:.3rem}
.kpi-lb{font-size:.7rem;font-weight:700;text-transform:uppercase;letter-spacing:.08em;color:var(--mu)}
.kpi-s{font-size:.72rem;color:var(--mu);margin-top:.2rem}

/* Downloads */
.dl-bar{background:var(--dk);border-radius:var(--rl);padding:1.3rem 1.6rem;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:.8rem;margin-bottom:1.8rem;box-shadow:var(--md)}
.dl-l{display:flex;align-items:center;gap:14px}
.dl-l .ico{font-size:1.5rem}
.dl-l h3{font-size:.92rem;font-weight:700;color:#fff}
.dl-l p{font-size:.73rem;color:rgba(255,255,255,.45);margin-top:1px}
.dl-btns{display:flex;gap:.6rem;flex-wrap:wrap}
.dl-btn{display:flex;align-items:center;gap:7px;padding:8px 15px;border-radius:var(--r);font-size:.75rem;font-weight:700;cursor:pointer;border:none;font-family:'Plus Jakarta Sans',sans-serif;transition:all .2s;text-decoration:none}
.dl-btn:hover{transform:translateY(-1px);filter:brightness(1.08)}
.dlP{background:linear-gradient(135deg,#0a1628,#0d2137);color:#fff;border:1px solid rgba(255,255,255,.15)}
.dlE{background:linear-gradient(135deg,#1a6b3c,#0e7c6e);color:#fff}
.dlS{background:linear-gradient(135deg,#c9a84c,#e8c45a);color:var(--dk)}
.dlA{background:linear-gradient(135deg,#d64045,#e07b39);color:#fff}
.dlF{background:linear-gradient(135deg,#457b9d,#7b5ea7);color:#fff}

/* Chart cards */
.ac{background:#fff;border-radius:var(--rx);border:1px solid var(--bd);box-shadow:var(--xs);overflow:hidden;transition:box-shadow .2s;margin-bottom:1.2rem}
.ac:hover{box-shadow:var(--sm)}
.ac-hdr{display:flex;align-items:center;justify-content:space-between;padding:1rem 1.4rem .8rem;border-bottom:1px solid var(--bd)}
.ac-hl{display:flex;align-items:center;gap:10px}
.ac-ic{font-size:1.1rem}
.ac-hdr h3{font-size:.9rem;font-weight:700;color:var(--dk)}
.ac-hdr p{font-size:.72rem;color:var(--mu);margin-top:1px}
.ac-hr{display:flex;align-items:center;gap:.5rem}
.ac-tag{font-size:.62rem;background:var(--lt);color:var(--gr);padding:3px 10px;border-radius:20px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;white-space:nowrap}
.btn-dlc{background:transparent;border:1px solid var(--bd);color:var(--mu);padding:4px 10px;border-radius:8px;font-size:.65rem;font-weight:700;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;transition:all .2s}
.btn-dlc:hover{background:var(--lt);color:var(--gr);border-color:var(--gr)}
.ac-img{width:100%;display:block;padding:.6rem .6rem .2rem}
.ch-ph{min-height:200px;display:flex;align-items:center;justify-content:center;color:var(--mu);font-size:.8rem;flex-direction:column;gap:.5rem;padding:2rem}
.ch-ph .phi{font-size:2rem;opacity:.3}
.ac-ins{margin:.2rem 1rem 1rem;background:linear-gradient(135deg,#f3fbf6,#edf7f2);border:1px solid rgba(26,107,60,.13);border-radius:var(--r);padding:1rem 1.2rem}
.ai-hdr{display:flex;align-items:center;gap:8px;margin-bottom:.6rem}
.ai-pill{font-size:.6rem;font-weight:700;text-transform:uppercase;letter-spacing:.1em;color:var(--gr);background:rgba(26,107,60,.1);padding:3px 10px;border-radius:20px}
.ac-ins p{font-size:.82rem;color:#253b2d;line-height:1.8;white-space:pre-wrap}
.g2{display:grid;grid-template-columns:1fr 1fr;gap:1.2rem}
@media(max-width:1100px){.g2{grid-template-columns:1fr}}

/* ML table */
.ml-tbl{background:#fff;border-radius:var(--rx);border:1px solid var(--bd);box-shadow:var(--xs);overflow:hidden;margin-bottom:1.2rem}
.ml-tbl table{width:100%;border-collapse:collapse;font-size:.83rem}
.ml-tbl thead{background:var(--dk)}
.ml-tbl thead th{padding:.85rem 1.2rem;text-align:left;font-size:.67rem;font-weight:700;text-transform:uppercase;letter-spacing:.08em;color:rgba(255,255,255,.65)}
.ml-tbl tbody tr{border-bottom:1px solid var(--bd);transition:background .15s}
.ml-tbl tbody tr:last-child{border:none}
.ml-tbl tbody tr:hover{background:#f8fbff}
.ml-tbl tbody td{padding:.8rem 1.2rem;color:var(--tx)}
.b-best{background:linear-gradient(135deg,var(--gr),var(--tl));color:#fff;font-size:.6rem;padding:2px 9px;border-radius:20px;font-weight:700;margin-left:8px}
.r2bar{display:flex;align-items:center;gap:10px}
.r2t{height:6px;width:80px;background:var(--lt);border-radius:3px;overflow:hidden;flex-shrink:0}
.r2f{height:100%;border-radius:3px;background:linear-gradient(90deg,var(--tl),var(--gr))}

/* Forecast table */
.fc-tbl{background:#fff;border-radius:var(--rx);border:1px solid var(--bd);box-shadow:var(--xs);overflow:hidden;margin-bottom:.8rem}
.fc-tbl table{width:100%;border-collapse:collapse;font-size:.83rem}
.fc-tbl thead{background:linear-gradient(135deg,#1a6b3c,#0e7c6e)}
.fc-tbl thead th{padding:.8rem 1.1rem;text-align:left;font-size:.67rem;font-weight:700;text-transform:uppercase;letter-spacing:.08em;color:rgba(255,255,255,.8)}
.fc-tbl tbody tr{border-bottom:1px solid var(--bd);transition:background .15s}
.fc-tbl tbody tr:last-child{border:none}
.fc-tbl tbody tr:hover{background:#f0faf5}
.fc-tbl tbody td{padding:.75rem 1.1rem;color:var(--tx)}
.cp{color:#16a34a;font-weight:700}.cn{color:var(--rd);font-weight:700}

/* Simulator */
.sim-card{background:#fff;border-radius:var(--rx);border:1px solid var(--bd);box-shadow:var(--xs);overflow:hidden}
.sim-body{padding:1.5rem}
.sim-g{display:grid;grid-template-columns:repeat(3,1fr);gap:1rem;margin-bottom:1.2rem}
@media(max-width:900px){.sim-g{grid-template-columns:repeat(2,1fr)}}
.sf{display:flex;flex-direction:column;gap:5px}
.sf label{font-size:.72rem;font-weight:700;color:var(--mu);text-transform:uppercase;letter-spacing:.07em}
.sf input,.sf select{border:1px solid var(--bd);border-radius:var(--r);padding:9px 12px;font-size:.88rem;font-family:'Plus Jakarta Sans',sans-serif;color:var(--tx);outline:none;transition:border-color .2s}
.sf input:focus,.sf select:focus{border-color:var(--gr);box-shadow:0 0 0 3px rgba(26,107,60,.1)}
.sim-res{background:linear-gradient(135deg,var(--dk),var(--nv));border-radius:var(--rl);padding:1.5rem 2rem;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:1rem;margin-top:1.2rem}
.sr-l h4{font-size:.72rem;font-weight:700;text-transform:uppercase;letter-spacing:.1em;color:rgba(255,255,255,.4);margin-bottom:.3rem}
.sr-v{font-family:'Fraunces',serif;font-size:2.8rem;font-weight:600;color:#fff;line-height:1}
.sr-u{font-size:.8rem;color:rgba(255,255,255,.5);margin-top:.2rem}
.sr-r{text-align:right}
.sr-rng{font-size:.75rem;color:rgba(255,255,255,.45);margin-bottom:.3rem}
.sr-ph{font-size:.9rem;color:var(--li);font-weight:700}
.btn-sim{background:linear-gradient(135deg,var(--gr),var(--tl));color:#fff;padding:11px 28px;border-radius:var(--r);font-size:.85rem;font-weight:700;border:none;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;transition:all .2s}
.btn-sim:hover{transform:translateY(-2px);box-shadow:0 4px 16px rgba(26,107,60,.35)}
.btn-sim:disabled{opacity:.5;pointer-events:none}

/* Estate modal */
#em{display:none;position:fixed;inset:0;z-index:1000;background:rgba(10,22,40,.7);backdrop-filter:blur(4px);overflow-y:auto}
#em.show{display:flex;align-items:flex-start;justify-content:center;padding:3rem 1rem}
.mb{background:#fff;border-radius:var(--rx);width:100%;max-width:900px;box-shadow:var(--lg);animation:up .3s ease;overflow:hidden}
.mh{background:linear-gradient(135deg,var(--dk),var(--nv));padding:1.4rem 1.8rem;display:flex;align-items:center;justify-content:space-between}
.mh h3{font-family:'Fraunces',serif;font-size:1.3rem;font-weight:600;color:#fff}
.mh p{font-size:.75rem;color:rgba(255,255,255,.5);margin-top:2px}
.mc{background:rgba(255,255,255,.1);border:none;color:rgba(255,255,255,.7);width:32px;height:32px;border-radius:8px;font-size:1.1rem;cursor:pointer;display:flex;align-items:center;justify-content:center;transition:all .2s;flex-shrink:0}
.mc:hover{background:rgba(255,255,255,.2);color:#fff}
.mbody{padding:1.5rem 1.8rem}
.mk-row{display:grid;grid-template-columns:repeat(4,1fr);gap:.8rem;margin-bottom:1.2rem}
.mk{background:var(--bg);border-radius:var(--r);padding:.8rem 1rem;text-align:center}
.mk-v{font-family:'Fraunces',serif;font-size:1.4rem;font-weight:600;color:var(--dk)}
.mk-l{font-size:.68rem;font-weight:700;color:var(--mu);text-transform:uppercase;letter-spacing:.07em;margin-top:.2rem}
.mi{width:100%;border-radius:var(--rl);margin-bottom:1rem}
.el{cursor:pointer;text-decoration:underline;text-underline-offset:2px;color:var(--gr)}
.el:hover{color:var(--tl)}

/* ── [NEW-9] CHAT WIDGET ── */
#chat-fab{
  position:fixed;bottom:28px;right:28px;z-index:900;
  width:58px;height:58px;border-radius:50%;
  background:linear-gradient(135deg,var(--gr),var(--tl));
  border:none;cursor:pointer;display:none;
  align-items:center;justify-content:center;font-size:1.5rem;
  box-shadow:0 6px 24px rgba(26,107,60,.45);transition:all .25s;
}
#chat-fab.hidden{display:none!important}
#chat-fab:hover{transform:scale(1.1);box-shadow:0 10px 32px rgba(26,107,60,.6)}
#chat-fab .fab-dot{
  position:absolute;top:-2px;right:-2px;width:18px;height:18px;
  background:var(--rd);border-radius:50%;font-size:.6rem;font-weight:700;
  color:#fff;display:flex;align-items:center;justify-content:center;border:2px solid #fff;
}
#chat-win{
  position:fixed;top:0;left:var(--sw);z-index:800;
  width:calc(100vw - var(--sw));height:100vh;background:#fff;
  border-radius:0;box-shadow:none;
  display:none;flex-direction:column;overflow:hidden;
  animation:slideIn .25s ease;border-left:1px solid var(--bd);
}
@keyframes slideIn{from{opacity:0}to{opacity:1}}
#chat-win.open{display:flex}

.ch-hdr{background:linear-gradient(135deg,var(--dk),var(--nv));padding:.9rem 1.2rem;display:flex;align-items:center;justify-content:space-between;flex-shrink:0}
.ch-av-wrap{display:flex;align-items:center;gap:10px}
.ch-av{width:36px;height:36px;border-radius:50%;background:linear-gradient(135deg,var(--gr),var(--tl));display:flex;align-items:center;justify-content:center;font-size:1rem;flex-shrink:0}
.ch-av-wrap h4{font-size:.9rem;font-weight:700;color:#fff}
.ch-av-wrap p{font-size:.68rem;color:rgba(255,255,255,.45);margin-top:1px}
.ch-online{width:8px;height:8px;border-radius:50%;background:var(--li);animation:blink 2s infinite;flex-shrink:0}
.btn-cc{background:rgba(255,255,255,.1);border:none;color:rgba(255,255,255,.7);width:28px;height:28px;border-radius:8px;cursor:pointer;font-size:.9rem;display:flex;align-items:center;justify-content:center;transition:all .2s;flex-shrink:0;font-family:'Plus Jakarta Sans',sans-serif}
.btn-cc:hover{background:rgba(255,255,255,.2);color:#fff}
.ch-body{flex:1;overflow-y:auto;padding:1.2rem;display:flex;flex-direction:column;gap:1rem;scroll-behavior:smooth}
.ch-body::-webkit-scrollbar{width:4px}
.ch-body::-webkit-scrollbar-thumb{background:var(--bd);border-radius:2px}
.msg{display:flex;gap:8px;max-width:90%;animation:up .2s ease}
.msg.ai{align-self:flex-start}
.msg.usr{align-self:flex-end;flex-direction:row-reverse}
.msg-av{width:28px;height:28px;border-radius:50%;flex-shrink:0;display:flex;align-items:center;justify-content:center;font-size:.8rem}
.msg.ai .msg-av{background:linear-gradient(135deg,var(--gr),var(--tl));color:#fff}
.msg.usr .msg-av{background:var(--lt);color:var(--gr);font-size:.75rem;font-weight:700}
.msg-bub{padding:.85rem 1.1rem;border-radius:12px;font-size:.85rem;line-height:1.75;max-width:340px}
.msg.ai .msg-bub{background:var(--bg);color:var(--tx);border-radius:4px 12px 12px 12px;border:1px solid var(--bd)}
.msg.usr .msg-bub{background:linear-gradient(135deg,var(--gr),var(--tl));color:#fff;border-radius:12px 4px 12px 12px}
.msg-t{font-size:.6rem;opacity:.45;margin-top:.3rem}
.msg.ai .msg-t{text-align:left}.msg.usr .msg-t{text-align:right}
.typing{display:flex;gap:4px;padding:.4rem .2rem}
.typing span{width:7px;height:7px;border-radius:50%;background:var(--mu);animation:bounce .9s ease-in-out infinite}
.typing span:nth-child(2){animation-delay:.15s}
.typing span:nth-child(3){animation-delay:.3s}
@keyframes bounce{0%,60%,100%{transform:translateY(0)}30%{transform:translateY(-6px)}}
.ch-qs{padding:.6rem .8rem;border-top:1px solid var(--bd);display:flex;flex-wrap:wrap;gap:.4rem;background:#f8fafc;flex-shrink:0}
.ch-qs .qs-lbl{font-size:.62rem;color:var(--mu);font-weight:700;text-transform:uppercase;letter-spacing:.07em;width:100%;margin-bottom:.1rem}
.cq{font-size:.72rem;background:#fff;border:1px solid var(--bd);color:var(--gr);padding:5px 10px;border-radius:20px;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;transition:all .18s;font-weight:600}
.cq:hover{background:var(--gr);color:#fff;border-color:var(--gr)}
.ch-inp{display:flex;align-items:center;gap:.5rem;padding:.8rem 1rem;border-top:1px solid var(--bd);background:#fff;flex-shrink:0}
.ch-inp input{flex:1;border:1px solid var(--bd);border-radius:20px;padding:8px 14px;font-size:.82rem;font-family:'Plus Jakarta Sans',sans-serif;outline:none;transition:border-color .2s;color:var(--tx)}
.ch-inp input:focus{border-color:var(--gr);box-shadow:0 0 0 3px rgba(26,107,60,.08)}
.btn-send{background:linear-gradient(135deg,var(--gr),var(--tl));border:none;color:#fff;width:34px;height:34px;border-radius:50%;cursor:pointer;display:flex;align-items:center;justify-content:center;font-size:.9rem;transition:all .2s;flex-shrink:0}
.btn-send:hover{transform:scale(1.1)}
.btn-send:disabled{opacity:.5;cursor:not-allowed}

/* Comparative */
.comp-c{background:#fff;border-radius:var(--rx);border:1px solid var(--bd);box-shadow:var(--xs);padding:1.5rem;margin-bottom:1rem}
.pr-a{font-size:.72rem;font-weight:700;padding:4px 12px;border-radius:20px;background:#dcfce7;color:#166534}
.pr-b{font-size:.72rem;font-weight:700;padding:4px 12px;border-radius:20px;background:#dbeafe;color:#1e40af}

.meta-r{background:#fff;border-radius:var(--rl);padding:.9rem 1.4rem;border:1px solid var(--bd);display:flex;justify-content:space-between;flex-wrap:wrap;gap:.5rem;font-size:.72rem;color:var(--mu);align-items:center;margin-top:1.5rem}
.meta-r strong{color:var(--dk)}

/* [FIX-11] Toast centered-bottom */
#toast{position:fixed;bottom:28px;left:50%;transform:translateX(-50%);background:var(--dk);color:#fff;padding:12px 22px;border-radius:var(--r);font-size:.8rem;font-weight:600;opacity:0;transition:all .3s;pointer-events:none;z-index:9998;border-left:3px solid var(--li);box-shadow:var(--lg);white-space:nowrap}
#toast.show{opacity:1;transform:translateX(-50%) translateY(-3px)}
#toast.err{border-left-color:var(--rd)}
::-webkit-scrollbar{width:5px}::-webkit-scrollbar-track{background:transparent}::-webkit-scrollbar-thumb{background:var(--bd);border-radius:3px}
@media(max-width:900px){#sb{transform:translateX(-100%)}#sb.open{transform:translateX(0)}#main{margin-left:0}}
</style>
</head>
<body>

<!-- LANDING -->
<div id="login-page" style="position:fixed;inset:0;z-index:99999;background:var(--dk);display:flex;align-items:center;justify-content:center">
  <div style="background:rgba(255,255,255,.05);border:1px solid rgba(255,255,255,.1);border-radius:20px;padding:2.5rem;width:100%;max-width:400px;box-shadow:var(--lg)">
    <div style="text-align:center;margin-bottom:2rem">
      <div style="width:56px;height:56px;border-radius:14px;background:linear-gradient(135deg,#1a6b3c,#0e7c6e);display:flex;align-items:center;justify-content:center;margin:0 auto 1rem;font-size:1.8rem">🌿</div>
      <div style="font-family:'Fraunces',serif;font-size:1.6rem;font-weight:600;color:#fff">LONSUM LEAP</div>
      <div style="font-size:.72rem;color:rgba(255,255,255,.4);margin-top:.3rem;text-transform:uppercase;letter-spacing:.1em">Plantation Intelligence v5.0</div>
    </div>
    <div id="login-err" style="display:none;background:rgba(214,64,69,.15);border:1px solid rgba(214,64,69,.3);color:#fca5a5;padding:10px 14px;border-radius:10px;font-size:.8rem;margin-bottom:1rem"></div>
    <div style="display:flex;flex-direction:column;gap:.9rem">
      <div>
        <label style="font-size:.7rem;font-weight:700;color:rgba(255,255,255,.45);text-transform:uppercase;letter-spacing:.08em;display:block;margin-bottom:.4rem">Username</label>
        <input id="l-user" type="text" placeholder="Masukkan username" style="width:100%;background:rgba(255,255,255,.07);border:1px solid rgba(255,255,255,.12);border-radius:10px;padding:11px 14px;font-size:.9rem;color:#fff;font-family:'Plus Jakarta Sans',sans-serif;outline:none" onkeydown="if(event.key==='Enter')doLogin()"/>
      </div>
      <div>
        <label style="font-size:.7rem;font-weight:700;color:rgba(255,255,255,.45);text-transform:uppercase;letter-spacing:.08em;display:block;margin-bottom:.4rem">Password</label>
        <input id="l-pass" type="password" placeholder="Masukkan password" style="width:100%;background:rgba(255,255,255,.07);border:1px solid rgba(255,255,255,.12);border-radius:10px;padding:11px 14px;font-size:.9rem;color:#fff;font-family:'Plus Jakarta Sans',sans-serif;outline:none" onkeydown="if(event.key==='Enter')doLogin()"/>
      </div>
      <button onclick="doLogin()" id="btn-login" style="background:linear-gradient(135deg,#1a6b3c,#0e7c6e);color:#fff;padding:13px;border-radius:10px;font-size:.9rem;font-weight:700;border:none;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;margin-top:.4rem;transition:all .2s">Masuk →</button>
    </div>
    <div style="text-align:center;margin-top:1.5rem;font-size:.68rem;color:rgba(255,255,255,.2)">PT London Sumatra Indonesia Tbk · Confidential</div>
  </div>
</div>

<div id="lp">
  <div style="position:fixed;inset:0;z-index:0"><div class="lp-mesh"></div><div class="lp-grid"></div>
    <div class="orb o1"></div><div class="orb o2"></div><div class="orb o3"></div></div>
  <div class="lp-wrap">
    <div class="lp-nav">
      <div class="lp-brand">
        <div class="lp-logo"><svg width="44" height="44" viewBox="0 0 44 44" fill="none"><rect width="44" height="44" rx="12" fill="url(#g1)"/><text x="21" y="30" text-anchor="middle" font-family="serif" font-size="21" font-weight="bold" fill="white">L</text><path d="M27 10 Q36 8 34 19 Q31 13 24 15 Z" fill="rgba(255,255,255,0.7)"/><defs><linearGradient id="g1" x1="0" y1="0" x2="44" y2="44"><stop offset="0%" stop-color="#1a6b3c"/><stop offset="100%" stop-color="#0e7c6e"/></linearGradient></defs></svg></div>
        <div><h1>PT London Sumatra Indonesia</h1><p>Lonsum · Tbk · Est. 1906</p></div>
      </div>
      <div style="display:flex;gap:.6rem"><span class="pill p-n">LEAP v5.0</span><span class="pill p-g">⚡ AI-Powered</span></div>
    </div>
    <div class="lp-main">
      <div class="eyebrow"><span class="dot-pulse"></span>Platform Analitik Perkebunan Enterprise — v5.0</div>
      <h1 class="lp-h1">LEAP<br/><span class="it">Plantation</span><span class="acc"> Intelligence</span></h1>
      <p class="lp-tag">Dashboard kecerdasan buatan untuk <strong>memantau, menganalisis, dan memprediksi</strong> produksi perkebunan Lonsum — kini dengan <strong>Chat with Your Data</strong>, PDF eksekutif, forecast 3 bulan, dan simulator What-If interaktif.</p>
      <button class="btn-start" onclick="enterApp()">Mulai Analisis <span class="arr">→</span></button>
      <div class="feat-grid">
        <div class="feat"><div class="feat-i">💬</div><div class="feat-t">Chat with Data <span class="new-badge">NEW</span></div><div class="feat-d">Tanya AI langsung tentang data Anda — produksi, tren, alert, rekomendasi</div></div>
        <div class="feat"><div class="feat-i">📄</div><div class="feat-t">PDF Annual Report</div><div class="feat-d">Cover page + 6 seksi + semua chart. Satu klik laporan eksekutif siap rapat direksi</div></div>
        <div class="feat"><div class="feat-i">📊</div><div class="feat-t">Perbandingan YoY</div><div class="feat-d">Bandingkan produksi 2 periode otomatis — upload CSV multi-tahun atau 2 file terpisah</div></div>
        <div class="feat"><div class="feat-i">🔮</div><div class="feat-t">Forecast 3 Bulan</div><div class="feat-d">Prediksi produksi dengan confidence interval yang melebar per horizon waktu</div></div>
        <div class="feat"><div class="feat-i">⚗️</div><div class="feat-t">What-If Simulator</div><div class="feat-d">"Jika curah hujan 250mm, 80 pekerja → produksi berapa?" — real-time</div></div>
        <div class="feat"><div class="feat-i">🔔</div><div class="feat-t">Alert Real-time</div><div class="feat-d">🔴 Kritis / 🟡 Perhatian / 🟢 Normal per estate langsung di dashboard</div></div>
        <div class="feat"><div class="feat-i">🏭</div><div class="feat-t">Estate Drilldown</div><div class="feat-d">Klik estate → tren historis, peringkat fleet, faktor dominan, AI insight</div></div>
        <div class="feat"><div class="feat-i">🖼️</div><div class="feat-t">Export Chart PNG</div><div class="feat-d">Download setiap grafik untuk PowerPoint dan presentasi eksekutif</div></div>
      </div>
      <div class="info-box">
        <div class="sec-lbl">Cara Penggunaan</div>
        <div class="steps">
          <div><div class="sn">1</div><div class="st">Siapkan CSV</div><div class="sd">Data produksi 7 kolom, atau 2 CSV untuk perbandingan periode</div></div>
          <div><div class="sn">2</div><div class="st">Upload &amp; Proses</div><div class="sd">Pilih mode Single/Comparative, upload, sistem proses ~30 detik</div></div>
          <div><div class="sn">3</div><div class="st">Explore Dashboard</div><div class="sd">Baca insight AI per grafik, klik estate, coba What-If simulator</div></div>
          <div><div class="sn">4</div><div class="st">Chat &amp; Unduh</div><div class="sd">Tanya AI tentang data Anda, export PDF atau Excel untuk arsip</div></div>
        </div>
      </div>
      <div class="fmt-box">
        <div class="fmt-in">
          <div class="sec-lbl">Kolom CSV yang Diperlukan</div>
          <div class="chips">
            <div class="chip"><code>date</code><span>YYYY-MM-DD</span></div>
            <div class="chip"><code>estate</code><span>Nama kebun</span></div>
            <div class="chip"><code>plantation_area_ha</code><span>Luas lahan (ha)</span></div>
            <div class="chip"><code>rainfall_mm</code><span>Curah hujan (mm)</span></div>
            <div class="chip"><code>workers</code><span>Jumlah tenaga kerja</span></div>
            <div class="chip"><code>fertilizer_kg</code><span>Pupuk (kg)</span></div>
            <div class="chip"><code>production_tons</code><span>Produksi (ton)</span></div>
          </div>
        </div>
      </div>
    </div>
    <div class="lp-foot">
      <p><strong>Lonsum LEAP v5.0</strong> — Enterprise Analytics Platform · Confidential</p>
      <div class="stags"><span class="stag">FastAPI</span><span class="stag">scikit-learn</span><span class="stag">Matplotlib</span><span class="stag">ReportLab</span><span class="stag">NVIDIA NIM</span><span class="stag">openpyxl</span></div>
    </div>
  </div>
</div>

<!-- SIDEBAR -->
<nav id="sb">
  <div class="sb-top">
    <div class="sb-logo">
      <div class="sb-li"><svg width="42" height="42" viewBox="0 0 42 42" fill="none"><rect width="42" height="42" rx="10" fill="url(#sbg)"/><text x="21" y="28" text-anchor="middle" font-family="serif" font-size="20" font-weight="bold" fill="white">L</text><path d="M26 12 Q32 10 30 18 Q28 14 22 15 Z" fill="rgba(255,255,255,0.7)"/><defs><linearGradient id="sbg" x1="0" y1="0" x2="42" y2="42"><stop offset="0%" stop-color="#1a6b3c"/><stop offset="100%" stop-color="#0e7c6e"/></linearGradient></defs></svg></div>
      <div class="sb-brand"><h1>LONSUM LEAP</h1><p>Intelligence Platform</p></div>
    </div>
    <div class="sb-mi"><span>⏰</span><span id="clk">—</span></div>
    <div class="sb-mi"><span>📍</span><span>PT London Sumatra Indonesia</span></div>
  </div>
  <div class="sb-nav">
    <div class="sb-sec">
      <button class="sb-btn" onclick="goHome()" style="border:1px solid rgba(255,255,255,.1);margin-bottom:.4rem"><span class="si">🏠</span>Kembali ke Beranda</button>
    </div>
    <div class="sb-sec">
      <span class="sb-lbl">Ringkasan</span>
      <button class="sb-btn on" onclick="nav('s-ov',this)"><span class="si">📊</span>Overview &amp; KPI</button>
      <button class="sb-btn" onclick="nav('s-al',this)"><span class="si">🔔</span>Alert Produksi</button>
      <button class="sb-btn" onclick="nav('s-dq',this)"><span class="si">✅</span>Data Quality</button>
      <button class="sb-btn" onclick="nav('s-dl',this)"><span class="si">📦</span>Unduh Laporan<span class="sb-bx">5</span></button>
    </div>
    <div class="sb-sec">
      <span class="sb-lbl">Analisis Produksi</span>
      <button class="sb-btn" onclick="nav('s-tr',this)"><span class="si">📈</span>Tren &amp; Musiman</button>
      <button class="sb-btn" onclick="nav('s-es',this)"><span class="si">🏭</span>Perbandingan Estate</button>
      <button class="sb-btn" onclick="nav('s-pr',this)"><span class="si">🌱</span>Produktivitas / Ha</button>
    </div>
    <div class="sb-sec">
      <span class="sb-lbl">Analisis Faktor</span>
      <button class="sb-btn" onclick="nav('s-co',this)"><span class="si">🔗</span>Korelasi &amp; Driver</button>
    </div>
    <div class="sb-sec">
      <span class="sb-lbl">Prediksi AI</span>
      <button class="sb-btn" onclick="nav('s-ml',this)"><span class="si">🤖</span>Performa Model ML</button>
      <button class="sb-btn" onclick="nav('s-fi',this)"><span class="si">🏆</span>Faktor Terpenting</button>
      <button class="sb-btn" onclick="nav('s-fc',this)"><span class="si">🔮</span>Forecast 3 Bulan</button>
      <button class="sb-btn" onclick="nav('s-si',this)"><span class="si">⚗️</span>What-If Simulator</button>
    </div>
    <div class="sb-sec">
      <span class="sb-lbl">Lanjutan</span>
      <button class="sb-btn" onclick="nav('s-cm',this)"><span class="si">📊</span>Perbandingan YoY</button>
      <button class="sb-btn" id="sb-chat" onclick="toggleChat()"><span class="si">💬</span>Chat with Data<span class="sb-bn">NEW</span></button>
    </div>
  </div>
  <div class="sb-bot">
    <button onclick="doLogout()" style="width:100%;background:rgba(214,64,69,.15);border:1px solid rgba(214,64,69,.25);color:#fca5a5;padding:9px;border-radius:var(--r);font-size:.78rem;font-weight:700;cursor:pointer;font-family:'Plus Jakarta Sans',sans-serif;margin-bottom:.6rem;transition:all .2s">🚪 Logout</button>
    <div class="sb-ver">LEAP v5.0 · scikit-learn · NVIDIA NIM · ReportLab</div>
  </div>
</nav>

<!-- MAIN -->
<div id="main">
  <div id="tb">
    <div class="tb-l">
      <h2 id="tb-t">Plantation Analytics</h2>
      <span class="tb-bc" id="tb-s">Upload CSV untuk memulai analisis</span>
    </div>
    <div class="tb-r">
      <span class="ai-badge">⚡ AI-Powered</span>
      <button id="btn-dm" onclick="toggleDark()" style="background:rgba(255,255,255,.08);border:1px solid rgba(255,255,255,.15);color:rgba(255,255,255,.7);width:36px;height:36px;border-radius:var(--r);font-size:1rem;cursor:pointer;transition:all .2s;display:flex;align-items:center;justify-content:center">🌙</button>
      <button class="btn-reset" id="btn-rst" onclick="resetDash()">↺ Upload Baru</button>
    </div>
  </div>

  <div id="cnt">

    <!-- UPLOAD -->
    <div id="up-sec">
      <div class="up-tabs">
        <button class="up-tab on" onclick="switchTab('single',this)">📁 Single Dataset</button>
      </div>
      <div class="up-pane on" id="pane-single">
        <div class="up-card" id="uz" onclick="document.getElementById('fi').click()"
            ondragover="onDrag(event)" ondragleave="this.classList.remove('dv')" ondrop="onDrop(event)">
          <div class="up-icon">🌿</div>
          <h2>Upload Data Produksi</h2>
          <p>Drag &amp; drop file CSV ke sini, atau klik untuk pilih file.<br/>
          <strong>Tip:</strong> Upload CSV multi-tahun → perbandingan YoY otomatis muncul!</p>
          <div class="col-hint">
            <code>date</code><code>estate</code><code>plantation_area_ha</code>
            <code>rainfall_mm</code><code>workers</code><code>fertilizer_kg</code><code>production_tons</code>
          </div>
          <button class="btn-p" onclick="event.stopPropagation();document.getElementById('fi').click()">📁 Pilih File CSV</button>
          <input type="file" id="fi" accept=".csv" onchange="onFile(event)"/>
        </div>
      </div>
      <div class="up-pane" id="pane-comp">
        <div style="background:#fff;border-radius:var(--rx);padding:2rem;border:1px solid var(--bd);box-shadow:var(--sm)">
          <div style="font-family:'Fraunces',serif;font-size:1.4rem;font-weight:600;color:var(--dk);margin-bottom:.3rem">Analisis Perbandingan Periode</div>
          <p style="font-size:.83rem;color:var(--mu);margin-bottom:1.2rem">Upload 2 CSV (2023 vs 2024) atau 1 CSV multi-tahun yang di-split otomatis.</p>
          <div class="cp2">
            <div class="cs" id="cs-a" onclick="document.getElementById('fia').click()"><div class="cs-i">📅</div><div class="cs-l">Periode A (Lebih Lama)</div><div class="cs-f" id="cs-an">Klik untuk upload</div><input type="file" id="fia" accept=".csv" onchange="onComp('a',event)"/></div>
            <div class="cs" id="cs-b" onclick="document.getElementById('fib').click()"><div class="cs-i">📅</div><div class="cs-l">Periode B (Lebih Baru)</div><div class="cs-f" id="cs-bn">Klik untuk upload</div><input type="file" id="fib" accept=".csv" onchange="onComp('b',event)"/></div>
          </div>
          <div style="text-align:center;margin:.8rem 0;font-size:.75rem;color:var(--mu)">— atau 1 CSV multi-tahun —</div>
          <div style="text-align:center;margin-bottom:1.2rem">
            <div class="cs" style="max-width:300px;display:inline-block" onclick="document.getElementById('fi').click()" id="cs-auto"><div class="cs-i">🔄</div><div class="cs-l">1 CSV Multi-Tahun (Auto Split)</div><div class="cs-f" id="cs-auton">Klik untuk upload</div></div>
          </div>
          <div style="text-align:center"><button class="btn-p" id="btn-comp" onclick="runComp()" disabled>📊 Jalankan Analisis Komparatif</button></div>
        </div>
      </div>
    </div>

    <!-- LOADING -->
    <div id="ld">
      <div class="sp-w"><div class="sp"></div><div class="spi"></div></div>
      <div id="ld-msg">Memulai pipeline analisis…</div>
      <div class="pb"><div class="pbf"></div></div>
      <div class="ls-wrap">
        <div class="ls" id="s1"><span>Membaca &amp; menilai kualitas data</span><span class="ck"></span></div>
        <div class="ls" id="s2"><span>Membuat grafik &amp; visualisasi</span><span class="ck"></span></div>
        <div class="ls" id="s3"><span>Melatih model prediksi ML</span><span class="ck"></span></div>
        <div class="ls" id="s4"><span>Menghasilkan 10 insight AI (paralel)</span><span class="ck"></span></div>
        <div class="ls" id="s5"><span>Menyiapkan laporan &amp; download</span><span class="ck"></span></div>
      </div>
    </div>
    <!-- DASHBOARD -->
    <div id="dash">
      <div id="s-al" style="margin-top:0;padding-top:0"><div id="al-wrap"></div></div>
      <div id="s-dq" style="margin-bottom:1.8rem;display:none">
        <div class="sep"><span class="sl">✅ Data Quality Score</span><div class="ln"></div></div>
        <div class="dq-row" id="dq-row"></div>
        <div class="ac">
          <div class="ac-hdr"><div class="ac-hl"><span class="ac-ic">✅</span><div><h3>Laporan Kualitas Data</h3><p>Completeness, outlier, dan duplikasi sebelum analisis</p></div></div><div class="ac-hr"><span class="ac-tag">Pre-Analysis</span><button class="btn-dlc" onclick="dlChart('c-dq','data_quality')">⬇ PNG</button></div></div>
          <img id="c-dq" alt="" class="ac-img" style="display:none"/><div class="ch-ph" id="c-dq-ph"><div class="phi">✅</div><span>Memuat…</span></div>
        </div>
      </div>

      <div id="s-ov">
        <div class="ph"><div class="ph-l">Executive Overview</div><h2>Ringkasan Performa Produksi</h2><p id="kpi-sub"></p></div>
        <div class="kpi-g" id="kpi-g"></div>
      </div>

      <div id="s-dl">
        <div class="sep"><span class="sl">📦 Unduh Laporan</span><div class="ln"></div></div>
        <div class="dl-bar">
          <div class="dl-l"><div class="ico">📂</div><div><h3>Laporan Otomatis Siap Diunduh</h3><p>PDF Annual Report + 4 Excel — format profesional siap rapat</p></div></div>
          <div class="dl-btns">
            <button class="dl-btn dlP" onclick="dlFile('/api/download/pdf','Lonsum_AnnualReport.pdf','application/pdf')">📄 Annual Report PDF</button>
            <button class="dl-btn dlE" onclick="dlFile('/api/download/excel','Lonsum_Produksi.xlsx')">📊 Produksi Excel</button>
            <button class="dl-btn dlS" onclick="dlFile('/api/download/stats','Lonsum_Statistik.xlsx')">📋 Statistik Estate</button>
            <button class="dl-btn dlA" onclick="dlFile('/api/download/alerts','Lonsum_Alert.xlsx')">⚠️ Alert</button>
            <button class="dl-btn dlF" onclick="dlFile('/api/download/forecast','Lonsum_Forecast.xlsx')">🔮 Forecast</button>
          </div>
        </div>
      </div>

      <div id="s-tr">
        <div class="sep"><span class="sl">📈 Tren &amp; Pola Musiman</span><div class="ln"></div></div>
        <div class="ac">
          <div class="ac-hdr"><div class="ac-hl"><span class="ac-ic">📈</span><div><h3>Tren Produksi Bulanan</h3><p>Total produksi per bulan + rolling avg 3 bulan</p></div></div><div class="ac-hr"><span class="ac-tag">Time Series</span><button class="btn-dlc" onclick="dlChart('c-tr','trend')">⬇ PNG</button></div></div>
          <img id="c-tr" alt="" class="ac-img" style="display:none"/><div class="ch-ph" id="c-tr-ph"><div class="phi">📈</div><span>Memuat…</span></div>
          <div class="ac-ins" id="ai-tr"><div class="ai-hdr"><span class="ai-pill">🤖 Insight AI</span></div><p>Menganalisis…</p></div>
        </div>
        <div class="g2">
          <div class="ac">
            <div class="ac-hdr"><div class="ac-hl"><span class="ac-ic">📅</span><div><h3>Profil Musiman</h3><p>Rata-rata produksi per bulan kalender</p></div></div><div class="ac-hr"><span class="ac-tag">Seasonality</span><button class="btn-dlc" onclick="dlChart('c-se','musiman')">⬇ PNG</button></div></div>
            <img id="c-se" alt="" class="ac-img" style="display:none"/><div class="ch-ph" id="c-se-ph"><div class="phi">📅</div><span>Memuat…</span></div>
            <div class="ac-ins" id="ai-se"><div class="ai-hdr"><span class="ai-pill">🤖 Insight AI</span></div><p>Menganalisis…</p></div>
          </div>
          <div class="ac" id="s-es">
            <div class="ac-hdr"><div class="ac-hl"><span class="ac-ic">🏭</span><div><h3>Produksi Tahunan per Estate</h3><p>Kontribusi setiap estate per tahun</p></div></div><div class="ac-hr"><span class="ac-tag">Stacked Bar</span><button class="btn-dlc" onclick="dlChart('c-an','tahunan')">⬇ PNG</button></div></div>
            <img id="c-an" alt="" class="ac-img" style="display:none"/><div class="ch-ph" id="c-an-ph"><div class="phi">🏭</div><span>Memuat…</span></div>
            <div class="ac-ins" id="ai-an"><div class="ai-hdr"><span class="ai-pill">🤖 Insight AI</span></div><p>Menganalisis…</p></div>
          </div>
        </div>
      </div>

      <div id="s-pr">
        <div class="sep"><span class="sl">🌱 Produktivitas &amp; Distribusi</span><div class="ln"></div></div>
        <div class="g2">
          <div class="ac">
            <div class="ac-hdr"><div class="ac-hl"><span class="ac-ic">📦</span><div><h3>Distribusi Produksi per Estate</h3><p>Boxplot — klik nama estate untuk drilldown</p></div></div><div class="ac-hr"><span class="ac-tag">Box Plot</span><button class="btn-dlc" onclick="dlChart('c-bp','boxplot')">⬇ PNG</button></div></div>
            <img id="c-bp" alt="" class="ac-img" style="display:none"/><div class="ch-ph" id="c-bp-ph"><div class="phi">📦</div><span>Memuat…</span></div>
            <div class="ac-ins" id="ai-bp"><div class="ai-hdr"><span class="ai-pill">🤖 Insight AI</span></div><p>Menganalisis…</p></div>
          </div>
          <div class="ac">
            <div class="ac-hdr"><div class="ac-hl"><span class="ac-ic">🌱</span><div><h3>Produktivitas per Hektar</h3><p>Klik estate di bawah untuk drilldown</p></div></div><div class="ac-hr"><span class="ac-tag">Benchmark</span><button class="btn-dlc" onclick="dlChart('c-ph','prodha')">⬇ PNG</button></div></div>
            <img id="c-ph" alt="" class="ac-img" style="display:none"/><div class="ch-ph" id="c-ph-ph"><div class="phi">🌱</div><span>Memuat…</span></div>
            <div class="ac-ins" id="ai-ph"><div class="ai-hdr"><span class="ai-pill">🤖 Insight AI</span></div><p>Menganalisis…</p></div>
          </div>
        </div>
        <div style="background:#fff;border-radius:var(--rl);padding:1rem 1.4rem;border:1px solid var(--bd);margin-bottom:1rem;display:flex;align-items:center;gap:1rem;flex-wrap:wrap">
          <span style="font-size:.72rem;font-weight:700;color:var(--mu);text-transform:uppercase;letter-spacing:.08em">🏭 Drilldown Estate →</span>
          <div id="est-btns" style="display:flex;flex-wrap:wrap;gap:.4rem"></div>
        </div>
      </div>

      <div id="s-co">
        <div class="sep"><span class="sl">🔗 Korelasi &amp; Driver</span><div class="ln"></div></div>
        <div class="g2">
          <div class="ac">
            <div class="ac-hdr"><div class="ac-hl"><span class="ac-ic">🔗</span><div><h3>Matriks Korelasi</h3><p>Kekuatan hubungan antar variabel</p></div></div><div class="ac-hr"><span class="ac-tag">Heatmap</span><button class="btn-dlc" onclick="dlChart('c-co','korelasi')">⬇ PNG</button></div></div>
            <img id="c-co" alt="" class="ac-img" style="display:none"/><div class="ch-ph" id="c-co-ph"><div class="phi">🔗</div><span>Memuat…</span></div>
            <div class="ac-ins" id="ai-co"><div class="ai-hdr"><span class="ai-pill">🤖 Insight AI</span></div><p>Menganalisis…</p></div>
          </div>
          <div class="ac">
            <div class="ac-hdr"><div class="ac-hl"><span class="ac-ic">⚙️</span><div><h3>Driver vs Produksi</h3><p>Scatter: input operasional vs output</p></div></div><div class="ac-hr"><span class="ac-tag">Scatter</span><button class="btn-dlc" onclick="dlChart('c-sc','scatter')">⬇ PNG</button></div></div>
            <img id="c-sc" alt="" class="ac-img" style="display:none"/><div class="ch-ph" id="c-sc-ph"><div class="phi">⚙️</div><span>Memuat…</span></div>
            <div class="ac-ins" id="ai-sc"><div class="ai-hdr"><span class="ai-pill">🤖 Insight AI</span></div><p>Menganalisis…</p></div>
          </div>
        </div>
      </div>

      <div id="s-ml">
        <div class="sep"><span class="sl">🤖 Model Machine Learning</span><div class="ln"></div></div>
        <div class="ml-tbl" style="margin-bottom:1.2rem">
          <table><thead><tr><th>Model</th><th>Akurasi (R²)</th><th>MAE (ton)</th><th>RMSE (ton)</th><th>CV R²</th><th>Status</th></tr></thead><tbody id="mtbl"></tbody></table>
        </div>
        <div class="ac">
          <div class="ac-hdr"><div class="ac-hl"><span class="ac-ic">🎯</span><div><h3>Evaluasi Model Terbaik</h3><p>Prediksi vs Aktual, Residual, Distribusi Residual</p></div></div><div class="ac-hr"><span class="ac-tag">Evaluation</span><button class="btn-dlc" onclick="dlChart('c-ml','model_eval')">⬇ PNG</button></div></div>
          <img id="c-ml" alt="" class="ac-img" style="display:none"/><div class="ch-ph" id="c-ml-ph"><div class="phi">🎯</div><span>Memuat…</span></div>
          <div class="ac-ins" id="ai-ml"><div class="ai-hdr"><span class="ai-pill">🤖 Insight AI</span></div><p>Menganalisis…</p></div>
        </div>
      </div>

      <div id="s-fi">
        <div class="sep"><span class="sl">🏆 Faktor Terpenting</span><div class="ln"></div></div>
        <div class="ac">
          <div class="ac-hdr"><div class="ac-hl"><span class="ac-ic">🏆</span><div><h3>Feature Importance</h3><p>Faktor paling menentukan hasil produksi</p></div></div><div class="ac-hr"><span class="ac-tag">Importance</span><button class="btn-dlc" onclick="dlChart('c-fi','feature_importance')">⬇ PNG</button></div></div>
          <img id="c-fi" alt="" class="ac-img" style="display:none"/><div class="ch-ph" id="c-fi-ph"><div class="phi">🏆</div><span>Memuat…</span></div>
          <div class="ac-ins" id="ai-fi"><div class="ai-hdr"><span class="ai-pill">🤖 Insight AI</span></div><p>Menganalisis…</p></div>
        </div>
      </div>

      <div id="s-fc">
        <div class="sep"><span class="sl">🔮 Forecast 3 Bulan</span><div class="ln"></div></div>
        <div class="ac">
          <div class="ac-hdr"><div class="ac-hl"><span class="ac-ic">🔮</span><div><h3>Prediksi 3 Bulan — Semua Estate</h3><p>CI melebar per horizon waktu</p></div></div><div class="ac-hr"><span class="ac-tag">3-Month</span><button class="btn-dlc" onclick="dlChart('c-fc','forecast')">⬇ PNG</button></div></div>
          <img id="c-fc" alt="" class="ac-img" style="display:none"/><div class="ch-ph" id="c-fc-ph"><div class="phi">🔮</div><span>Memuat…</span></div>
        </div>
        <div class="fc-tbl" id="fc-tbl" style="display:none">
          <table><thead><tr><th>Estate</th><th>Bulan +1 (ton)</th><th>Bulan +2 (ton)</th><th>Bulan +3 (ton)</th><th>Aktual Terakhir</th><th>Tren 3 Bulan</th></tr></thead><tbody id="fc-body"></tbody></table>
        </div>
        <div class="ac-ins" id="ai-fc" style="margin-top:.5rem"><div class="ai-hdr"><span class="ai-pill">🤖 Insight AI</span></div><p>Menganalisis…</p></div>
      </div>

      <div id="s-si">
        <div class="sep"><span class="sl">⚗️ What-If Simulator</span><div class="ln"></div></div>
        <div class="sim-card">
          <div class="ac-hdr"><div class="ac-hl"><span class="ac-ic">⚗️</span><div><h3>What-If Scenario Simulator</h3><p>Input kondisi → prediksi real-time dari model ML terlatih</p></div></div><span class="ac-tag">Interactive ML</span></div>
          <div class="sim-body">
            <div class="sim-g">
              <div class="sf"><label>Estate</label><select id="sim-est"></select></div>
              <div class="sf"><label>Bulan Target</label><select id="sim-mo"><option value="1">Januari</option><option value="2">Februari</option><option value="3">Maret</option><option value="4">April</option><option value="5">Mei</option><option value="6">Juni</option><option value="7">Juli</option><option value="8">Agustus</option><option value="9">September</option><option value="10">Oktober</option><option value="11">November</option><option value="12">Desember</option></select></div>
              <div class="sf"><label>Luas Lahan (ha)</label><input type="number" id="sim-a" placeholder="contoh: 500" min="1"/></div>
              <div class="sf"><label>Curah Hujan (mm)</label><input type="number" id="sim-r" placeholder="contoh: 200" min="0"/></div>
              <div class="sf"><label>Jumlah Pekerja</label><input type="number" id="sim-w" placeholder="contoh: 80" min="1"/></div>
              <div class="sf"><label>Pupuk (kg)</label><input type="number" id="sim-f" placeholder="contoh: 5000" min="0"/></div>
            </div>
            <button class="btn-sim" id="btn-sim" onclick="runSim()">⚗️ Hitung Prediksi</button>
            <div class="sim-res" id="sim-res" style="display:none">
              <div class="sr-l"><h4>Prediksi Produksi</h4><div class="sr-v" id="sr-v">—</div><div class="sr-u">ton / bulan</div></div>
              <div class="sr-r"><div class="sr-rng" id="sr-rng"></div><div class="sr-ph" id="sr-ph"></div><div style="font-size:.72rem;color:rgba(255,255,255,.4);margin-top:.3rem" id="sr-vs"></div></div>
            </div>
          </div>
        </div>
      </div>

      <div id="s-cm">
        <div class="sep"><span class="sl">📊 Analisis Perbandingan Periode</span><div class="ln"></div></div>
        <div id="comp-cnt"><div style="background:#fff;border-radius:var(--rx);padding:2rem;border:1px solid var(--bd);text-align:center"><div style="font-size:2rem;margin-bottom:.8rem">📊</div><div style="font-size:.9rem;color:var(--mu)">Gunakan tab <strong>Comparative</strong> di bagian upload untuk analisis perbandingan periode.</div></div></div>
      </div>

      <div class="meta-r">
        <span>Dibuat: <strong id="meta-d"></strong></span>
        <span style="font-family:'JetBrains Mono',monospace;font-size:.68rem">PT London Sumatra Indonesia · LEAP v5.0</span>
        <span id="meta-m"></span>
      </div>
    </div><!-- /dash -->
  </div><!-- /cnt -->
</div><!-- /main -->

<!-- ESTATE MODAL -->
<div id="em">
  <div class="mb">
    <div class="mh"><div><h3 id="em-n">Estate Detail</h3><p id="em-s">Analisis mendalam performa estate</p></div><button class="mc" onclick="closeModal()">✕</button></div>
    <div class="mbody">
      <div class="mk-row" id="mk-row"></div>
      <img id="m-img" alt="" class="mi" style="display:none"/>
      <div class="ac-ins" id="m-ins"><div class="ai-hdr"><span class="ai-pill">🤖 Insight AI — Estate Spesifik</span></div><p>Memuat…</p></div>
    </div>
  </div>
</div>

<!-- CHAT FAB & WINDOW -->
<button id="chat-fab" onclick="toggleChat()" title="Chat with Your Data">
  💬
  <div class="fab-dot" id="fab-dot">1</div>
</button>

<div id="chat-win">
  <div class="ch-hdr">
    <div class="ch-av-wrap">
      <div class="ch-av">🤖</div>
      <div><h4>AI Data Analyst</h4><p>Berbasis data produksi Anda</p></div>
      <div class="ch-online"></div>
    </div>
    <button class="btn-cc" onclick="toggleChat()">✕</button>
  </div>
  <div class="ch-body" id="ch-body">
    <div class="msg ai">
      <div class="msg-av">🤖</div>
      <div>
        <div class="msg-bub">Halo! Saya AI analyst Lonsum LEAP v5.0. Setelah Anda upload data, saya bisa menjawab pertanyaan tentang produksi, tren, estate, forecast, dan rekomendasi operasional. 🌿</div>
        <div class="msg-t">sekarang</div>
      </div>
    </div>
  </div>
  <div class="ch-qs" id="ch-qs" style="display:none">
    <span class="qs-lbl">Pertanyaan cepat:</span>
    <button class="cq" onclick="sendQ('Estate mana yang produktivitasnya paling tinggi?')">🏆 Estate terbaik?</button>
    <button class="cq" onclick="sendQ('Apa faktor yang paling mempengaruhi produksi?')">🔬 Faktor utama?</button>
    <button class="cq" onclick="sendQ('Berikan ringkasan kondisi produksi saat ini')">📊 Ringkasan</button>
    <button class="cq" onclick="sendQ('Estate mana yang perlu perhatian khusus?')">⚠️ Alert estate?</button>
    <button class="cq" onclick="sendQ('Bagaimana prediksi produksi bulan depan?')">🔮 Forecast?</button>
    <button class="cq" onclick="sendQ('Apa rekomendasi utama untuk meningkatkan produksi?')">💡 Rekomendasi?</button>
  </div>
  <div class="ch-inp">
    <input type="text" id="ch-in" placeholder="Tanya tentang data produksi Anda…" onkeydown="if(event.key==='Enter')sendChat()"/>
    <button class="btn-send" id="btn-send" onclick="sendChat()">↑</button>
  </div>
</div>

<div id="toast"></div>

<script>
var _ch={},_kpis={},_estates=[],_mae=0;
var _chatReady=false,_chatHist=[];
var _token=null;

async function doLogin(){
  var btn=document.getElementById('btn-login');
  var err=document.getElementById('login-err');
  var u=document.getElementById('l-user').value.trim();
  var p=document.getElementById('l-pass').value.trim();
  if(!u||!p){err.style.display='';err.textContent='Username dan password wajib diisi.';return;}
  btn.textContent='Memverifikasi...';btn.disabled=true;err.style.display='none';
  var fd=new FormData();fd.append('username',u);fd.append('password',p);
  try{
    var r=await fetch('/api/auth/login',{method:'POST',body:fd});
    var d=await r.json();
    if(!r.ok){err.style.display='';err.textContent=d.detail||'Login gagal.';btn.textContent='Masuk';btn.disabled=false;return;}
    _token=d.access_token;
    document.getElementById('login-page').style.display='none';
    document.querySelector('.sb-mi span:last-child').textContent=d.full_name;
    btn.textContent='Masuk';btn.disabled=false;
  }catch(e){
    err.style.display='';err.textContent='Koneksi error.';btn.textContent='Masuk';btn.disabled=false;
  }
}
function doLogout(){
  _token=null;_chatReady=false;_chatHist=[];
  resetDash();
  document.getElementById('login-page').style.display='flex';
  document.getElementById('l-user').value='';
  document.getElementById('l-pass').value='';
  document.getElementById('login-err').style.display='none';
}

function toggleDark(){
  document.body.classList.toggle('dark');
  var btn=document.getElementById('btn-dm');
  btn.textContent=document.body.classList.contains('dark')?'☀️':'🌙';
  localStorage.setItem('leap-dark',document.body.classList.contains('dark')?'1':'0');
}

// Restore dark mode preference saat load
(function(){
  if(localStorage.getItem('leap-dark')==='1'){
    document.body.classList.add('dark');
    var btn=document.getElementById('btn-dm');
    if(btn)btn.textContent='☀️';
  }
})();

(function tick(){
  var d=new Date();
  var s=d.toLocaleDateString('id-ID',{weekday:'short',day:'2-digit',month:'short',year:'numeric'})+' · '+d.toLocaleTimeString('id-ID',{hour:'2-digit',minute:'2-digit',second:'2-digit'});
  var e=document.getElementById('clk');if(e)e.textContent=s;setTimeout(tick,1000);
})();

function nav(id,btn){
  var el=document.getElementById(id);if(!el)return;
  var th=62;window.scrollTo({top:el.getBoundingClientRect().top+window.scrollY-th-20,behavior:'smooth'});
  document.querySelectorAll('.sb-btn').forEach(function(b){b.classList.remove('on')});
  if(btn)btn.classList.add('on');
}
function toast(msg,type){
  var t=document.getElementById('toast');t.textContent=msg;
  t.className='show'+(type?' '+type:'');setTimeout(function(){t.className=''},4500);
}
function esc(s){if(s==null)return'—';return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}
function fmt(v){var n=parseFloat(v);return isNaN(n)?String(v==null?'—':v):n.toLocaleString('id-ID',{maximumFractionDigits:2});}
function setImg(id,b64){
  var img=document.getElementById(id),ph=document.getElementById(id+'-ph');
  if(!img)return;
  if(b64&&b64.length>200){img.onload=function(){img.style.display='block';if(ph)ph.style.display='none';};img.src='data:image/png;base64,'+b64;}
  else{img.style.display='none';if(ph){ph.querySelector('span').textContent='Data tidak tersedia.';ph.style.display='flex';}}
}
function setIns(id,text){
  var box=document.getElementById(id);if(!box)return;
  box.innerHTML='<div class="ai-hdr"><span class="ai-pill">🤖 Insight AI Analyst</span></div><p>'+esc(text&&text.length>10?text:'Insight AI tidak tersedia.')+'</p>';
}
function dlChart(imgId,name){
  var img=document.getElementById(imgId);
  if(!img||!img.src||img.style.display==='none'){toast('Grafik belum tersedia.','err');return;}
  var a=document.createElement('a');a.href=img.src;a.download='Lonsum_'+name+'_'+new Date().toISOString().slice(0,10)+'.png';a.click();toast('📥 '+name+'.png diunduh');
}

function dlFile(url, filename, mime){
  if(!_token){toast('Session expired, silakan login ulang.','err');return;}
  toast('⏳ Menyiapkan '+filename+'…');
  var urlWithToken = url + '?token=' + encodeURIComponent(_token);
  fetch(urlWithToken)
  .then(function(r){
    if(r.status===401){toast('Session expired, silakan login ulang.','err');doLogout();throw new Error('401');}
    if(!r.ok) throw new Error('Server error '+r.status);
    return r.blob();
  })
  .then(function(blob){
    var a=document.createElement('a');
    a.href=URL.createObjectURL(blob);
    a.download=filename;
    document.body.appendChild(a);
    a.click();
    setTimeout(function(){URL.revokeObjectURL(a.href);document.body.removeChild(a);},1000);
    toast('✅ '+filename+' berhasil diunduh');
  })
  .catch(function(e){
    if(e.message!=='401') toast('❌ Gagal unduh: '+e.message,'err');
  });
}

function onDrag(e){e.preventDefault();document.getElementById('uz').classList.add('dv');}
function onDrop(e){e.preventDefault();document.getElementById('uz').classList.remove('dv');var f=e.dataTransfer.files[0];if(f)doUpload(f);}
function onFile(e){var f=e.target.files[0];if(!f)return;doUpload(f);}

function showErr(msg){
  var z=document.getElementById('uz');var old=z.querySelector('.err-b');if(old)old.remove();
  var d=document.createElement('div');d.className='err-b';d.textContent='⚠️ '+msg;z.insertBefore(d,z.firstChild);
}
function animSteps(){
  var ids=['s1','s2','s3','s4','s5'],del=[0,8000,18000,28000,36000];
  ids.forEach(function(id,i){
    setTimeout(function(){
      ids.forEach(function(x,j){
        var el=document.getElementById(x);
        var ck=el.querySelector('.ck');
        if(j<i){
          el.className='ls done';
          if(ck)ck.textContent='✓';
        }else if(j===i){
          el.className='ls on';
          if(ck)ck.textContent='';
        }else{
          el.className='ls';
          if(ck)ck.textContent='';
        }
      });
    },del[i]);
  });
}
function clearSteps(){
  ['s1','s2','s3','s4','s5'].forEach(function(id){
    var el=document.getElementById(id);
    el.className='ls';
    var ck=el.querySelector('.ck');
    if(ck)ck.textContent='';
  });
}

function doUpload(file){
  if(!file.name.toLowerCase().endsWith('.csv')){toast('Hanya file CSV.','err');return;}
  var old=document.querySelector('#uz .err-b');if(old)old.remove();
  document.getElementById('up-sec').style.display='none';
  document.getElementById('ld').style.display='flex';
  document.getElementById('dash').style.display='none';
  clearSteps();animSteps();
  document.getElementById('tb-s').textContent='Menganalisis data…';
  var fd=new FormData();fd.append('file',file);
  fetch('/api/analyze',{method:'POST',body:fd,headers:{'Authorization':'Bearer '+_token}})
    .then(function(res){var st=res.status;return res.text().then(function(t){return{st:st,t:t};});})
    .then(function(o){
      if(o.st!==200){var em='Server error '+o.st;try{var p=JSON.parse(o.t);em=p.detail||em;}catch(e){}throw new Error(em);}
      var data;try{data=JSON.parse(o.t);}catch(e){throw new Error('Response parse gagal.');}
      renderDash(data);
    })
    .catch(function(err){clearSteps();document.getElementById('ld').style.display='none';document.getElementById('up-sec').style.display='';showErr(err.message);toast(err.message,'err');});
}

function runComp(){
  var fd=new FormData();
  if(_cMode==='pair'){fd.append('file_a',_cfA);fd.append('file_b',_cfB);fd.append('mode','pair');}
  else if(_cMode==='auto'&&_cfAuto){fd.append('file',_cfAuto);fd.append('mode','auto');}
  else{toast('Pilih file terlebih dahulu.','err');return;}
  document.getElementById('up-sec').style.display='none';document.getElementById('ld').style.display='flex';
  clearSteps();animSteps();
  fetch('/api/analyze/comparative',{method:'POST',body:fd,headers:{'Authorization':'Bearer '+_token}})
    .then(function(r){return r.json();}).then(function(data){renderComp(data);})
    .catch(function(err){clearSteps();document.getElementById('ld').style.display='none';document.getElementById('up-sec').style.display='';toast(err.message,'err');});
}

function renderDash(data){
  try{
    var k=data.kpis||{};_kpis=k;_estates=k.estates||[];_mae=data.model_results&&data.model_results[0]?data.model_results[0].mae:0;
    _ch=data.charts||{};_chatReady=true;
    document.getElementById('tb-t').textContent='Dashboard Produksi';
    document.getElementById('tb-s').textContent='Periode: '+(k.date_range||'—')+' · '+(k.num_estates||0)+' Estate';
    document.getElementById('kpi-sub').textContent='Periode: '+(k.date_range||'—')+'  ·  '+(k.num_estates||0)+' Estate  ·  '+fmt(k.total_records)+' Record';

    var cards=[
      {i:'🌿',l:'Total Produksi',v:fmt(k.total_production_tons)+' ton',s:'Seluruh estate',c:'ac-g'},
      {i:'📐',l:'Produktivitas Avg',v:fmt(k.avg_productivity_t_ha)+' t/ha',s:'Ton per hektar',c:'ac-go'},
      {i:'🏆',l:'Estate Terbaik',v:esc(k.best_estate||'—'),s:'Produksi tertinggi',c:'ac-t'},
      {i:'📅',l:'Bulan Puncak',v:esc(k.peak_month||'—'),s:'Rata-rata tertinggi',c:'ac-g'},
      {i:'🏭',l:'Jumlah Estate',v:k.num_estates||0,s:'Kebun dipantau',c:'ac-t'},
      {i:'📋',l:'Total Record',v:fmt(k.total_records),s:'Data diproses',c:'ac-r'},
    ];
    document.getElementById('kpi-g').innerHTML=cards.map(function(c){
      return '<div class="kpi"><div class="kpi-ac '+c.c+'"></div><div class="kpi-top"><div class="kpi-iw">'+c.i+'</div></div><div class="kpi-v">'+c.v+'</div><div class="kpi-lb">'+c.l+'</div><div class="kpi-s">'+c.s+'</div></div>';
    }).join('');

    var alerts=data.alert_data||[];
    document.getElementById('al-wrap').innerHTML=alerts.length?alerts.map(function(a){
      var cls=a.level==='crit'?'crit':a.level==='warn'?'warn':'ok';
      var ico=a.level==='crit'?'🔴':a.level==='warn'?'🟡':'🟢';
      return '<div class="al '+cls+'"><div class="al-ico">'+ico+'</div><div class="al-b"><h4>'+esc(a.estate)+' — '+esc(a.level_label)+'</h4><p>'+esc(a.message)+'</p></div></div>';
    }).join(''):'';

    var dq=data.data_quality||{};
    if(dq.score!==undefined){
      document.getElementById('s-dq').style.display='';
      var gc=dq.score>=80?'#1a6b3c':dq.score>=60?'#c9a84c':'#d64045';
      document.getElementById('dq-row').innerHTML=[
        {l:'Completeness',v:dq.completeness+'%',c:dq.completeness>95?'#1a6b3c':dq.completeness>85?'#c9a84c':'#d64045',w:dq.completeness},
        {l:'Outlier Rate',v:dq.outlier_rate+'%',c:dq.outlier_rate<5?'#1a6b3c':dq.outlier_rate<15?'#c9a84c':'#d64045',w:Math.max(0,100-dq.outlier_rate*3)},
        {l:'Duplikat',v:dq.duplicate_count,c:dq.duplicate_count===0?'#1a6b3c':'#e07b39',w:dq.duplicate_count===0?100:70},
        {l:'Overall Score',v:dq.score+'/100',c:gc,w:dq.score},
      ].map(function(c){return '<div class="dq-c"><div class="dq-v" style="color:'+c.c+'">'+c.v+'</div><div class="dq-l">'+c.l+'</div><div class="dq-tr"><div class="dq-tf" style="width:'+c.w+'%;background:'+c.c+'"></div></div></div>';}).join('');
      setImg('c-dq',_ch.dq);
    }

    setImg('c-tr',_ch.trend);setImg('c-se',_ch.seasonal);setImg('c-an',_ch.annual);
    setImg('c-bp',_ch.boxplot);setImg('c-ph',_ch.prodha);setImg('c-co',_ch.corr);
    setImg('c-sc',_ch.scatter);setImg('c-ml',_ch.model_eval);setImg('c-fi',_ch.feature_imp);setImg('c-fc',_ch.forecast);

    var ai=data.ai_insights||{};
    setIns('ai-tr',ai.trend);setIns('ai-se',ai.seasonal);setIns('ai-an',ai.annual);
    setIns('ai-bp',ai.boxplot);setIns('ai-ph',ai.prodha);setIns('ai-co',ai.correlation);
    setIns('ai-sc',ai.scatter);setIns('ai-ml',ai.model);setIns('ai-fi',ai.feature_importance);setIns('ai-fc',ai.forecast);

    var best=data.best_model||'';
    document.getElementById('mtbl').innerHTML=(data.model_results||[]).map(function(m){
      var r2=m.r2||0,pct=(r2*100).toFixed(1),bw=Math.round(Math.max(0,Math.min(1,r2))*80),iB=m.model===best;
      return '<tr style="'+(iB?'background:#f0faf5;':'')+'"><td><strong>'+esc(m.model)+'</strong>'+(iB?'<span class="b-best">★ Terbaik</span>':'')+'</td>'+
        '<td><div class="r2bar"><div class="r2t"><div class="r2f" style="width:'+bw+'px"></div></div><span style="font-weight:700;color:'+(r2>0.85?'#1a6b3c':r2>0.6?'#c9a84c':'#d64045')+'">'+pct+'%</span></div></td>'+
        '<td>'+fmt(m.mae)+' ton</td><td>'+fmt(m.rmse)+' ton</td><td>'+((m.cv_r2||0)*100).toFixed(1)+'%</td>'+
        '<td><span style="font-size:.7rem;padding:3px 10px;border-radius:20px;background:'+(iB?'#d1fae5':'#f1f5f9')+';color:'+(iB?'#065f46':'#64748b')+';font-weight:700;">'+(iB?'✓ Dipilih':'Dievaluasi')+'</span></td></tr>';
    }).join('');

    var fc=data.forecast_3m||[];
    if(fc.length>0){
      document.getElementById('fc-tbl').style.display='';
      document.getElementById('fc-body').innerHTML=fc.map(function(r){
        var chg=parseFloat(r.chg_m3||0),tr=chg>5?'📈 Naik':chg<-5?'📉 Turun':'➡ Stabil';
        return '<tr><td><span class="el" onclick="openModal(\''+esc(r.estate)+'\')">'+esc(r.estate)+'</span></td>'+
          '<td><strong>'+fmt(r.m1)+'</strong></td><td>'+fmt(r.m2)+'</td><td>'+fmt(r.m3)+'</td>'+
          '<td>'+fmt(r.last_actual)+'</td><td class="'+(chg>=0?'cp':'cn')+'">'+tr+' ('+Math.abs(chg).toFixed(1)+'%)</td></tr>';
      }).join('');
    }

    document.getElementById('est-btns').innerHTML=_estates.map(function(e){
      return '<button onclick="openModal(\''+esc(e)+'\')" style="background:var(--lt);border:1px solid rgba(26,107,60,.2);color:var(--gr);padding:5px 14px;border-radius:20px;font-size:.75rem;font-weight:700;cursor:pointer;font-family:\'Plus Jakarta Sans\',sans-serif;transition:all .2s" onmouseover="this.style.background=\'var(--gr)\';this.style.color=\'#fff\'" onmouseout="this.style.background=\'var(--lt)\';this.style.color=\'var(--gr)\'">🏭 '+esc(e)+'</button>';
    }).join('');

    document.getElementById('sim-est').innerHTML=_estates.map(function(e){return'<option value="'+esc(e)+'">'+esc(e)+'</option>';}).join('');
    if(data.estate_stats&&_estates.length>0){
      var fs=data.estate_stats[_estates[0]]||{};
      if(fs.avg_area)document.getElementById('sim-a').value=Math.round(fs.avg_area);
      if(fs.avg_rainfall)document.getElementById('sim-r').value=Math.round(fs.avg_rainfall);
      if(fs.avg_workers)document.getElementById('sim-w').value=Math.round(fs.avg_workers);
      if(fs.avg_fertilizer)document.getElementById('sim-f').value=Math.round(fs.avg_fertilizer);
    }

    document.getElementById('meta-d').textContent=data.generated_at||'—';
    document.getElementById('meta-m').textContent='Model: '+esc(best);
    clearSteps();
    document.getElementById('ld').style.display='none';document.getElementById('dash').style.display='block';
    document.getElementById('btn-rst').style.display='inline-flex';
    document.getElementById('chat-fab').style.display='flex';
    document.getElementById('ch-qs').style.display='';
    window.scrollTo({top:0,behavior:'smooth'});
    // Auto render comparative kalau ada
    if(data.comparative){
      renderComp(data.comparative);
    }
    toast('Dashboard siap — '+Object.keys(_ch).length+' grafik · '+_estates.length+' estate');
  }catch(err){
    console.error('[renderDash]',err);clearSteps();document.getElementById('ld').style.display='none';
    document.getElementById('up-sec').style.display='';showErr('Render error: '+err.message);toast('Render error','err');
  }
}

function renderComp(data){
  clearSteps();document.getElementById('ld').style.display='none';
  document.getElementById('dash').style.display='block';document.getElementById('btn-rst').style.display='inline-flex';
  var cc=document.getElementById('comp-cnt');
  if(!data||!data.summary){cc.innerHTML='<p style="color:var(--mu)">Data tidak tersedia.</p>';return;}
  var s=data.summary;
  cc.innerHTML='<div class="comp-c">'+
    '<div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:1.2rem">'+
    '<div style="font-family:\'Fraunces\',serif;font-size:1.1rem;font-weight:600;color:var(--dk)">'+esc(s.period_a)+' vs '+esc(s.period_b)+'</div>'+
    '<div style="display:flex;gap:.4rem"><span class="pr-a">'+esc(s.period_a)+'</span><span style="color:var(--mu);font-size:.8rem;font-weight:600">VS</span><span class="pr-b">'+esc(s.period_b)+'</span></div></div>'+
    '<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:1rem;margin-bottom:1rem">'+
    '<div style="background:var(--bg);border-radius:var(--r);padding:1rem;text-align:center"><div style="font-family:\'Fraunces\',serif;font-size:1.5rem;font-weight:600;color:var(--dk)">'+fmt(s.total_a)+' ton</div><div style="font-size:.7rem;color:var(--mu);font-weight:700;text-transform:uppercase;margin-top:.2rem">'+esc(s.period_a)+'</div></div>'+
    '<div style="background:var(--bg);border-radius:var(--r);padding:1rem;text-align:center"><div style="font-family:\'Fraunces\',serif;font-size:1.5rem;font-weight:600;color:var(--dk)">'+fmt(s.total_b)+' ton</div><div style="font-size:.7rem;color:var(--mu);font-weight:700;text-transform:uppercase;margin-top:.2rem">'+esc(s.period_b)+'</div></div>'+
    '<div style="background:'+(s.change_pct>=0?'#f0fdf4':'#fff1f0')+';border-radius:var(--r);padding:1rem;text-align:center">'+
    '<div style="font-family:\'Fraunces\',serif;font-size:1.5rem;font-weight:600;color:'+(s.change_pct>=0?'#166534':'#991b1b')+'">'+(s.change_pct>=0?'▲':'▼')+Math.abs(s.change_pct).toFixed(1)+'%</div>'+
    '<div style="font-size:.7rem;color:var(--mu);font-weight:700;text-transform:uppercase;margin-top:.2rem">Perubahan YoY</div></div></div>'+
    (data.charts&&data.charts.comparative?'<img src="data:image/png;base64,'+data.charts.comparative+'" style="width:100%;border-radius:var(--rl);margin-bottom:1rem"/>':'')+
    '<div class="ac-ins"><div class="ai-hdr"><span class="ai-pill">🤖 Insight AI</span></div><p>'+esc(data.ai_insight||'—')+'</p></div></div>';
  toast('Perbandingan YoY tersedia di menu Perbandingan YoY');
}

function runSim(){
  var btn=document.getElementById('btn-sim');btn.disabled=true;btn.textContent='⏳ Menghitung…';
  var p={estate:document.getElementById('sim-est').value,month:parseInt(document.getElementById('sim-mo').value),
         area_ha:parseFloat(document.getElementById('sim-a').value)||0,
         rainfall_mm:parseFloat(document.getElementById('sim-r').value)||0,
         workers:parseInt(document.getElementById('sim-w').value)||0,
         fertilizer_kg:parseFloat(document.getElementById('sim-f').value)||0};
  fetch('/api/predict',{method:'POST',headers:{'Content-Type':'application/json','Authorization':'Bearer '+_token},body:JSON.stringify(p)})
    .then(function(r){return r.json();})
    .then(function(d){
      document.getElementById('sim-res').style.display='flex';
      document.getElementById('sr-v').textContent=fmt(d.prediction)+' ton';
      document.getElementById('sr-rng').textContent='Rentang: '+fmt(d.lower)+' — '+fmt(d.upper)+' ton';
      var pha=p.area_ha>0?(d.prediction/p.area_ha).toFixed(4):0;
      document.getElementById('sr-ph').textContent=pha+' t/ha';
      var diff=(pha-(_kpis.avg_productivity_t_ha||0));
      document.getElementById('sr-vs').textContent=(diff>=0?'▲':'▼')+Math.abs(diff).toFixed(4)+' t/ha vs fleet avg';
      btn.disabled=false;btn.textContent='⚗️ Hitung Prediksi';
    })
    .catch(function(e){toast('Simulator error: '+e.message,'err');btn.disabled=false;btn.textContent='⚗️ Hitung Prediksi';});
}

function openModal(estate){
  document.getElementById('em-n').textContent=estate;
  document.getElementById('em-s').textContent='Analisis mendalam — '+estate;
  document.getElementById('em').classList.add('show');
  document.getElementById('m-img').style.display='none';
  document.getElementById('m-ins').innerHTML='<div class="ai-hdr"><span class="ai-pill">🤖 Insight AI</span></div><p>Memuat analisis…</p>';
  fetch('/api/estate/'+encodeURIComponent(estate),{headers:{'Authorization':'Bearer '+_token}})
    .then(function(r){return r.json();})
    .then(function(d){
      document.getElementById('mk-row').innerHTML=[
        {l:'Total Produksi',v:fmt(d.total_production)+' ton'},
        {l:'Avg Bulanan',v:fmt(d.avg_monthly)+' ton'},
        {l:'Produktivitas',v:d.avg_productivity+' t/ha'},
        {l:'Peringkat Fleet',v:'#'+d.fleet_rank+' / '+d.fleet_total},
      ].map(function(k){return'<div class="mk"><div class="mk-v">'+k.v+'</div><div class="mk-l">'+k.l+'</div></div>';}).join('');
      if(d.chart){var mc=document.getElementById('m-img');mc.src='data:image/png;base64,'+d.chart;mc.style.display='block';}
      document.getElementById('m-ins').innerHTML='<div class="ai-hdr"><span class="ai-pill">🤖 Insight AI — '+esc(estate)+'</span></div><p>'+esc(d.ai_insight)+'</p>';
    })
    .catch(function(){document.getElementById('m-ins').innerHTML='<div class="ai-hdr"><span class="ai-pill">🤖 Insight AI</span></div><p>Gagal memuat analisis estate.</p>';});
}
function closeModal(){document.getElementById('em').classList.remove('show');}
document.getElementById('em').addEventListener('click',function(e){if(e.target===this)closeModal();});

/* ══ CHAT WITH YOUR DATA ══════════════════════════════ */
function toggleChat(){
  var win=document.getElementById('chat-win');
  var main=document.getElementById('main');
  win.classList.toggle('open');
  main.classList.toggle('chat-open');
  if(win.classList.contains('open')){
    document.getElementById('fab-dot').style.display='none';
    document.getElementById('ch-in').focus();
    scrollChat();
  }
}
function scrollChat(){var b=document.getElementById('ch-body');b.scrollTop=b.scrollHeight;}
function addMsg(role,text){
  var body=document.getElementById('ch-body');
  var now=new Date().toLocaleTimeString('id-ID',{hour:'2-digit',minute:'2-digit'});
  var av=role==='ai'?'🤖':'U';
  body.innerHTML+='<div class="msg '+role+'"><div class="msg-av">'+av+'</div><div><div class="msg-bub">'+esc(text)+'</div><div class="msg-t">'+now+'</div></div></div>';
  scrollChat();
}
function addTyping(){
  var body=document.getElementById('ch-body');
  var div=document.createElement('div');div.className='msg ai';div.id='typing-msg';
  div.innerHTML='<div class="msg-av">🤖</div><div><div class="msg-bub"><div class="typing"><span></span><span></span><span></span></div></div></div>';
  body.appendChild(div);scrollChat();
}
function removeTyping(){var t=document.getElementById('typing-msg');if(t)t.remove();}

function sendQ(q){document.getElementById('ch-in').value=q;sendChat();}
function sendChat(){
  var inp=document.getElementById('ch-in');var q=inp.value.trim();if(!q)return;
  if(!_chatReady){addMsg('ai','⚠️ Upload data terlebih dahulu agar saya bisa menjawab pertanyaan tentang dataset Anda.');inp.value='';return;}
  inp.value='';addMsg('usr',q);
  _chatHist.push({role:'user',content:q});
  document.getElementById('btn-send').disabled=true;
  addTyping();
  fetch('/api/chat',{method:'POST',headers:{'Content-Type':'application/json','Authorization':'Bearer '+_token},
    body:JSON.stringify({message:q,history:_chatHist.slice(-8)})})
    .then(function(r){return r.json();})
    .then(function(d){
      removeTyping();var reply=d.reply||'Maaf, terjadi kesalahan.';
      addMsg('ai',reply);_chatHist.push({role:'assistant',content:reply});
      document.getElementById('btn-send').disabled=false;
    })
    .catch(function(e){removeTyping();addMsg('ai','⚠️ Koneksi error: '+e.message);document.getElementById('btn-send').disabled=false;});
}

function enterApp(){
  var lp=document.getElementById('lp');lp.classList.add('out');
  setTimeout(function(){lp.style.display='none';},620);
  // Reset state saat masuk app
  document.getElementById('dash').style.display='none';
  document.getElementById('up-sec').style.display='';
  document.getElementById('ld').style.display='none';
  document.getElementById('btn-rst').style.display='none';
}
function goHome(){
  // Bersihkan chat history saat kembali ke beranda
  _chatReady=false;
  _chatHist=[];
  document.getElementById('ch-body').innerHTML=
    '<div class="msg ai"><div class="msg-av">🤖</div><div>'+
    '<div class="msg-bub">Halo! Saya AI analyst Lonsum LEAP v5.0. Setelah Anda upload data, '+
    'saya bisa menjawab pertanyaan tentang produksi, tren, estate, forecast, dan rekomendasi operasional. 🌿</div>'+
    '<div class="msg-t">sekarang</div></div></div>';
  document.getElementById('ch-qs').style.display='none';
  document.getElementById('chat-win').classList.remove('open');
  document.getElementById('main').classList.remove('chat-open');
  document.getElementById('fab-dot').style.display='flex';
  document.getElementById('chat-fab').style.display='none';

  var lp=document.getElementById('lp');
  lp.style.display='';
  lp.classList.remove('out');
  window.scrollTo({top:0,behavior:'smooth'});
}
function resetDash(){
  document.getElementById('up-sec').style.display='';document.getElementById('dash').style.display='none';
  document.getElementById('btn-rst').style.display='none';document.getElementById('fi').value='';
  document.getElementById('chat-fab').style.display='none';document.getElementById('chat-win').classList.remove('open');
  _chatReady=false;_chatHist=[];
  document.getElementById('tb-t').textContent='Plantation Analytics';document.getElementById('tb-s').textContent='Upload CSV untuk memulai analisis';
  document.getElementById('ch-body').innerHTML='<div class="msg ai"><div class="msg-av">🤖</div><div><div class="msg-bub">Halo! Saya AI analyst Lonsum LEAP v5.0. Setelah Anda upload data, saya bisa menjawab pertanyaan tentang produksi, tren, estate, forecast, dan rekomendasi operasional. 🌿</div><div class="msg-t">sekarang</div></div></div>';
  document.getElementById('ch-qs').style.display='none';
  document.getElementById('fab-dot').style.display='flex';

  
  window.scrollTo({top:0,behavior:'smooth'});
}
document.addEventListener('keydown',function(e){if(e.key==='Escape'){closeModal();if(document.getElementById('chat-win').classList.contains('open'))toggleChat();}});
// Dynamic sidebar highlight on scroll
window.addEventListener('scroll', function(){
  if(document.getElementById('dash').style.display==='none') return;
  var sections=['s-al','s-dq','s-ov','s-dl','s-tr','s-es','s-pr','s-co','s-ml','s-fi','s-fc','s-si','s-cm'];
  var navMap={
    's-al':'s-al','s-dq':'s-dq','s-ov':'s-ov','s-dl':'s-dl',
    's-tr':'s-tr','s-es':'s-es','s-pr':'s-pr','s-co':'s-co',
    's-ml':'s-ml','s-fi':'s-fi','s-fc':'s-fc','s-si':'s-si','s-cm':'s-cm'
  };
  var scrollY=window.scrollY+80;
  var active=null;
  sections.forEach(function(id){
    var el=document.getElementById(id);
    if(!el||el.style.display==='none') return;
    if(el.getBoundingClientRect().top+window.scrollY<=scrollY) active=id;
  });
  if(!active) return;
  document.querySelectorAll('.sb-btn').forEach(function(b){b.classList.remove('on')});
  document.querySelectorAll('.sb-btn').forEach(function(b){
    if(b.getAttribute('onclick')&&b.getAttribute('onclick').includes("'"+active+"'"))
      b.classList.add('on');
  });
},{ passive:true });
</script>
</body>
</html>"""


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ROUTES
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
@app.get("/", response_class=HTMLResponse)
async def root(): return HTMLResponse(HTML_PAGE)

@app.post("/api/auth/login")
async def login(form: OAuth2PasswordRequestForm = Depends()):
    user = USERS_DB.get(form.username)
    if not user or not verify_password(form.password, user["hashed_password"]):
        raise HTTPException(401, "Username atau password salah")
    token = create_token({"sub": user["username"], "role": user["role"]})
    return {
        "access_token": token,
        "token_type": "bearer",
        "full_name": user["full_name"],
        "role": user["role"]
    }

@app.get("/api/auth/me")
async def me(user = Depends(get_current_user)):
    return {"username": user["username"], "full_name": user["full_name"], "role": user["role"]}

@app.post("/api/analyze")
async def analyze(file: UploadFile = File(...), user = Depends(get_current_user)):
    if not file.filename.lower().endswith(".csv"):
        raise HTTPException(400, "Only CSV files accepted.")
    try: df_raw=pd.read_csv(io.StringIO((await file.read()).decode("utf-8")))
    except Exception as e: raise HTTPException(400, f"Read CSV error: {e}")
    miss={"date","estate","plantation_area_ha","rainfall_mm","workers","fertilizer_kg","production_tons"}-set(df_raw.columns)
    if miss: raise HTTPException(422, f"Missing columns: {', '.join(sorted(miss))}")
    try: result=process_dataset(df_raw)
    except Exception as e:
        import traceback; traceback.print_exc(); raise HTTPException(500, f"Processing error: {e}")
    jb=json.dumps(result,ensure_ascii=False).encode("utf-8")
    return StreamingResponse(iter([jb]),media_type="application/json",headers={"Content-Length":str(len(jb))})


@app.post("/api/analyze/comparative")
async def analyze_comp(mode:str=Form(...),file:UploadFile=File(None),
                       file_a:UploadFile=File(None),file_b:UploadFile=File(None), user = Depends(get_current_user)):
    try:
        if mode=="auto" and file:
            df=pd.read_csv(io.StringIO((await file.read()).decode("utf-8")))
            df["date"]=pd.to_datetime(df["date"]); years=sorted(df["date"].dt.year.unique())
            if len(years)<2: raise HTTPException(422,"Butuh minimal 2 tahun untuk mode auto.")
            dfa=df[df["date"].dt.year==years[-2]].copy(); dfb=df[df["date"].dt.year==years[-1]].copy()
            la,lb=str(years[-2]),str(years[-1])
        elif mode=="pair" and file_a and file_b:
            dfa=pd.read_csv(io.StringIO((await file_a.read()).decode("utf-8")))
            dfb=pd.read_csv(io.StringIO((await file_b.read()).decode("utf-8")))
            dfa["date"]=pd.to_datetime(dfa["date"]); dfb["date"]=pd.to_datetime(dfb["date"])
            ya=dfa["date"].dt.year.mode()[0]; yb=dfb["date"].dt.year.mode()[0]
            la=str(ya); lb=str(yb) if ya!=yb else f"Periode B ({yb})"
        else: raise HTTPException(400,"Invalid mode atau file tidak ada.")
        return compute_comparative(dfa,dfb,str(la),str(lb))
    except HTTPException: raise
    except Exception as e:
        import traceback; traceback.print_exc(); raise HTTPException(500,f"Comparative error: {e}")


@app.post("/api/predict")
async def predict(payload: dict, user = Depends(get_current_user)):
    if "_best_mdl" not in _last: raise HTTPException(404,"Upload CSV terlebih dahulu.")
    try:
        mdl=_last["_best_mdl"]; le=_last["_le"]; FEAT=_last["_FEAT"]; mae=_last["_mae"]
        est=payload.get("estate","")
        if est not in le.classes_: raise HTTPException(422,f"Estate '{est}' tidak ditemukan.")
        month=int(payload.get("month",1))
        fv=pd.DataFrame([{"plantation_area_ha":float(payload.get("area_ha",0)),
                          "rainfall_mm":float(payload.get("rainfall_mm",0)),
                          "workers":int(payload.get("workers",0)),
                          "fertilizer_kg":float(payload.get("fertilizer_kg",0)),
                          "month":month,"quarter":(month-1)//3+1,
                          "estate_encoded":int(le.transform([est])[0])}])
        pred=float(mdl.predict(fv[FEAT])[0])
        return {"prediction":round(pred,2),"lower":round(pred-mae,2),"upper":round(pred+mae,2)}
    except HTTPException: raise
    except Exception as e: raise HTTPException(500,f"Predict error: {e}")


@app.get("/api/estate/{name}")
async def estate_detail(name: str, user = Depends(get_current_user)):
    if "_df" not in _last: raise HTTPException(404,"Data belum tersedia.")
    try:
        d=get_estate_detail(name)
        if not d: raise HTTPException(404,f"Estate '{name}' tidak ditemukan.")
        return d
    except HTTPException: raise
    except Exception as e:
        import traceback; traceback.print_exc(); raise HTTPException(500,f"Estate error: {e}")


# [NEW-9] Chat with Your Data endpoint
@app.post("/api/chat")
async def chat(payload: dict, user = Depends(get_current_user)):
    if "_chat_ctx" not in _last:
        return {"reply":"Belum ada data yang diupload. Silakan upload CSV terlebih dahulu, kemudian saya bisa menjawab pertanyaan tentang data produksi Anda."}
    try:
        ctx=_last["_chat_ctx"]; msg=payload.get("message","").strip(); hist=payload.get("history",[])
        SYS_CHAT=(
            "Anda adalah AI analyst perkebunan PT London Sumatra Indonesia (Lonsum). "
            "Jawab berdasarkan DATA AKTUAL berikut:\n\n"+ctx+"\n\n"
            "Panduan: Bahasa Indonesia formal, sertakan angka spesifik, singkat (maks 4 paragraf), "
            "jika pertanyaan di luar data produksi jelaskan batasannya dengan sopan. "
            "Akhiri dengan insight/rekomendasi singkat bila relevan."
        )
        messages=[{"role":"system","content":SYS_CHAT}]
        for h in hist[-8:]: messages.append({"role":h.get("role","user"),"content":h.get("content","")})
        messages.append({"role":"user","content":msg})
        body={"model":NVIDIA_MODEL,"messages":messages,"temperature":0.5,"max_tokens":600}
        hdrs={"Authorization":f"Bearer {NVIDIA_API_KEY}","Content-Type":"application/json"}
        with httpx.Client(timeout=65.0) as c:
            r=c.post(NVIDIA_BASE_URL,json=body,headers=hdrs); r.raise_for_status()
            reply=r.json()["choices"][0]["message"]["content"].strip()
        return {"reply":reply}
    except Exception as e:
        return {"reply":f"Maaf, terjadi kesalahan: {type(e).__name__}. Coba ulangi pertanyaan Anda."}


@app.get("/api/download/pdf")
async def download_pdf(request: Request, token: str = None, user = Depends(get_current_user)):
    if "_df" not in _last: raise HTTPException(404,"Belum ada data.")
    try:
        pdf=build_pdf(_last["kpis"],_last["model_results"],_last["forecast_3m"],
                      _last["alert_data"],_last["charts"],_last["ai_insights"])
    except Exception as e:
        import traceback; traceback.print_exc(); raise HTTPException(500,f"PDF error: {e}")
    fname=f"Lonsum_AnnualReport_{datetime.now().strftime('%Y%m%d')}.pdf"
    return Response(content=pdf,media_type="application/pdf",
                    headers={"Content-Disposition":f'attachment; filename="{fname}"'})

@app.get("/api/download/excel")
async def download_excel(request: Request, token: str = None, user = Depends(get_current_user)):
    if "_df" not in _last: raise HTTPException(404,"Belum ada data.")
    try:
        df=_last["_df"]; kpis=_last["kpis"]
        gen=kpis.get("generated_at","—"); dr=kpis.get("date_range","—")
        wb=Workbook(); ws=wb.active; ws.title="Laporan Produksi Bulanan"; ws.sheet_properties.tabColor="1a6b3c"
        _title(ws,"PT LONDON SUMATRA INDONESIA — LAPORAN PRODUKSI BULANAN",8,"1a6b3c")
        _sub(ws,f"Periode: {dr}  |  Dibuat: {gen}",8)
        for i,h in enumerate(["Tahun","Bulan","Estate","Luas (ha)","Curah Hujan (mm)","Tenaga Kerja","Pupuk (kg)","Produksi (ton)"],1): ws.cell(4,i,h)
        _hdr(ws,4,8,"1a6b3c")
        md=df.sort_values(["year","month","estate"]).reset_index(drop=True)
        for ri,row in md.iterrows():
            r=ri+5
            for ci,v in enumerate([int(row["year"]),str(row["month_name"]),str(row["estate"]),
                                    round(float(row["plantation_area_ha"]),2),round(float(row["rainfall_mm"]),1),
                                    int(row["workers"]),round(float(row["fertilizer_kg"]),1),
                                    round(float(row["production_tons"]),2)],1): ws.cell(r,ci,v)
            _drow(ws,r,8,alt=(ri%2==0))
        _widths(ws,[8,10,20,14,18,14,14,18])
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    except Exception as e: import traceback; traceback.print_exc(); raise HTTPException(500,f"Excel error: {e}")
    return Response(content=buf.read(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition":'attachment; filename="Lonsum_ProduksiBulanan.xlsx"'})


@app.get("/api/download/stats")
async def download_stats(request: Request, token: str = None, user = Depends(get_current_user)):
    if "_df" not in _last: raise HTTPException(404,"Belum ada data.")
    try:
        df=_last["_df"]; kpis=_last["kpis"]
        gen=kpis.get("generated_at","—"); dr=kpis.get("date_range","—")
        st=df.groupby("estate").agg(
            total=("production_tons","sum"),avg=("production_tons","mean"),
            mx=("production_tons","max"),mn=("production_tons","min"),
            std=("production_tons","std"),ph=("productivity_ton_per_ha","mean"),
            rain=("rainfall_mm","mean"),wk=("workers","mean"),
            cnt=("production_tons","count")).round(3).reset_index()
        wb=Workbook(); ws=wb.active; ws.title="Statistik Estate"; ws.sheet_properties.tabColor="0e7c6e"
        _title(ws,"STATISTIK ESTATE — PT LONDON SUMATRA INDONESIA",10,"0e7c6e")
        _sub(ws,f"Dibuat: {gen} | Periode: {dr}",10)
        for i,h in enumerate(["Estate","Total (ton)","Avg Bulanan","Maks","Min","Std Dev","Prod/ha","Hujan mm","Pekerja","Record"],1): ws.cell(4,i,h)
        _hdr(ws,4,10,"0e7c6e")
        for ri,row in st.iterrows():
            r=ri+5
            for ci,v in enumerate([row["estate"],row["total"],row["avg"],row["mx"],row["mn"],
                                    row["std"],row["ph"],row["rain"],
                                    round(float(row["wk"]),0),int(row["cnt"])],1): ws.cell(r,ci,v)
            _drow(ws,r,10,alt=(ri%2==0))
        _widths(ws,[22,16,16,14,14,12,14,14,12,10])
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    except Exception as e: import traceback; traceback.print_exc(); raise HTTPException(500,f"Stats error: {e}")
    return Response(content=buf.read(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition":'attachment; filename="Lonsum_StatistikEstate.xlsx"'})


@app.get("/api/download/alerts")
async def download_alerts(request: Request, token: str = None, user = Depends(get_current_user)):
    if "_alerts_df" not in _last: raise HTTPException(404,"Belum ada data.")
    try:
        df=_last["_df"]; al=_last["_alerts_df"].copy(); kpis=_last["kpis"]
        gen=kpis.get("generated_at","—")
        avg_ph=float(df["productivity_ton_per_ha"].mean()); thr=avg_ph*.75
        al["fleet_avg"]=round(avg_ph,4)
        al["deficit_pct"]=((avg_ph-al["productivity_ton_per_ha"])/avg_ph*100).round(1)
        wb=Workbook(); ws=wb.active; ws.title="Alert Produktivitas"; ws.sheet_properties.tabColor="d64045"
        _title(ws,"ALERT PRODUKTIVITAS — PT LONDON SUMATRA INDONESIA",7,"d64045")
        _sub(ws,f"Threshold: <{thr:.3f} t/ha  |  Fleet avg: {avg_ph:.3f} t/ha  |  Dibuat: {gen}",7)
        for i,h in enumerate(["Tanggal","Estate","Produksi (ton)","Luas (ha)","Prod/ha","Fleet Avg","Defisit (%)"],1): ws.cell(4,i,h)
        _hdr(ws,4,7,"d64045")
        if len(al)>0:
            for ri,(_,row) in enumerate(al.iterrows()):
                r=ri+5; pv=float(row["productivity_ton_per_ha"]); def_=float(row["deficit_pct"])
                try: ds=row["date"].strftime("%b %Y")
                except: ds=str(row["date"])
                for ci,v in enumerate([ds,str(row["estate"]),round(float(row["production_tons"]),2),
                                        round(float(row["plantation_area_ha"]),2),
                                        round(pv,4),round(avg_ph,4),round(def_,1)],1): ws.cell(r,ci,v)
                _drow(ws,r,7,alt=(ri%2==0))
                if def_>40:
                    for c in range(1,8): ws.cell(r,c).fill=_hfill("FEE2E2")
        _widths(ws,[14,20,16,14,16,16,14])
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    except Exception as e: import traceback; traceback.print_exc(); raise HTTPException(500,f"Alerts error: {e}")
    return Response(content=buf.read(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition":'attachment; filename="Lonsum_AlertProduktivitas.xlsx"'})


@app.get("/api/download/forecast")
async def download_forecast(request: Request, token: str = None, user = Depends(get_current_user)):
    if "_fc_df" not in _last: raise HTTPException(404,"Belum ada data.")
    try:
        fc=_last["_fc_df"].copy(); kpis=_last["kpis"]
        gen=kpis.get("generated_at","—"); best=_last.get("best_model","—")
        wb=Workbook(); ws=wb.active; ws.title="Forecast 3 Bulan"; ws.sheet_properties.tabColor="c9a84c"
        _title(ws,"FORECAST PRODUKSI 3 BULAN — PT LONDON SUMATRA INDONESIA",8,"c9a84c")
        _sub(ws,f"Model: {best}  |  Dibuat: {gen}",8)
        for i,h in enumerate(["Estate","Bulan +1 (ton)","Bulan +2 (ton)","Bulan +3 (ton)","Aktual Terakhir","Tren M1 (%)","Tren M3 (%)"],1): ws.cell(4,i,h)
        _hdr(ws,4,8,"c9a84c")
        for ri,row in enumerate(fc.itertuples()):
            r=ri+5
            for ci,v in enumerate([row.estate,round(row.m1,2),round(row.m2,2),round(row.m3,2),
                                    round(row.last_actual,2),
                                    f"{row.chg_m1:+.1f}%",f"{row.chg_m3:+.1f}%"],1): ws.cell(r,ci,v)
            _drow(ws,r,8,alt=(ri%2==0))
        _widths(ws,[22,16,16,16,18,14,14])
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    except Exception as e: import traceback; traceback.print_exc(); raise HTTPException(500,f"Forecast error: {e}")
    return Response(content=buf.read(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition":'attachment; filename="Lonsum_Forecast3Bulan.xlsx"'})


@app.get("/api/health")
async def health():
    return {"status":"ok","has_data":"_df" in _last,"version":"5.0.0","timestamp":datetime.now().isoformat()}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=False)
